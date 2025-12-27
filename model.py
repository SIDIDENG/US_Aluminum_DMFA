'''Metalscape'''
### Version 10/30/2025
### 10/30/2025: Added self.primary_consumed and self.sweetener_consumed
### 10/24/2025: Overhaul Scenario A and B
### 07/22/2025: Breaking down p.CO2_oprt into p.CO2_dom_to_dom + p.CO2_export
### 07/03/2025: Added detailed tracking for trade scenarios within system_deets
### 06/30/2025: Implemented dynamic programming to def CO2_ctg
### 06/25/2025: update_A: dynamically adjust A.loc['E','R1] and A.loc['E','R2]
### 06/20/2025: Added SC_profile dataframe
### 05/17/2025: Added import product emissions
### 05/09/2025: Export alloy breakdowns to excel
### 05/07/2025: Add semi_matrix
### 04/04/2025: monitoring inert anodes in system_deets
### 03/26/2025: non-negative Sink columns
### 03/10/2025: Made the upper scrap import limits as system arguments
### 03/08/2025: set target R1 post-consumer scrap ratio for the improving recycling scenario
### 02/07/2025: adjusted PFC and inert anode settings
### 01/12/2024: overhauled EoL scrap generation
### 01/08/2025: fixed a few dtype issues
### 12/13/2024: added SA
### 09/12/2024: added primary high domestic setting, adjusted A matrix


import simpy
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border

from Pre_processing import *
pd.set_option('display.max_columns', 10)
#%%
### MFA unit: kilo Tonnes/year

# In[]
class g:    
    process = pd.read_excel("inputs.xlsx", sheet_name = 'Process',usecols = 'A:L,O,R',nrows = 24)
    process.set_index("Label", inplace = True)
    
    A = pd.read_excel("inputs.xlsx", sheet_name = 'A', usecols = 'A:W', nrows =  22)
    A.set_index('A Matrix', inplace = True)
    A.fillna(0, inplace = True)

    Sink = pd.read_excel("inputs.xlsx", sheet_name = 'A', usecols = 'A,X:AB',nrows =  22)
    Sink.set_index('A Matrix', inplace = True)
    
    fab_efficiencies = pd.read_excel("inputs.xlsx", sheet_name = "downstream",usecols= 'N:V', skiprows = 24, nrows = 7)
    fab_efficiencies.set_index('Process',inplace = True)
    
    product_trade_table = pd.read_excel("inputs.xlsx", sheet_name = "end_use",usecols= 'A,M:R', nrows = 8)
    product_trade_table['import_ratio'] = product_trade_table['import']/product_trade_table['dom_fabricated']
    product_trade_table['export_ratio'] = product_trade_table['export']/product_trade_table['dom_consumed']
    product_trade_table['dom_to_dom_ratio'] = (product_trade_table['dom_fabricated'] - product_trade_table['export'])/product_trade_table['dom_consumed']
    product_trade_table.set_index("idx", inplace = True)
    ### List of post-consumer scrap other than container
    production_idx_list = list(product_trade_table.index)
    production_idx_list.remove('P7')
    
    EoL_table = pd.read_excel("inputs.xlsx", sheet_name = "upstream", usecols= 'L,M', skiprows = 0, nrows=5)
    EoL_table.set_index('index', inplace = True)
    EoL_table['ratio'] = EoL_table['amount']/EoL_table.loc['collected','amount']
    
    consumption_target = generate_consumption_target_init(A, Sink)
    IO_matrix = generate_IO_matrix_init(A, Sink)
    flow = generate_flow(IO_matrix,process)
    
    semi_CF = pd.read_excel("inputs.xlsx", sheet_name = 'semi_CF', usecols = 'A:F', nrows = 7)    
    semi_CF.set_index('Label', inplace = True)
    semi_to_finished = pd.read_excel("inputs.xlsx", sheet_name = 'semi_CF', usecols = 'J,L:S', nrows = 7)   
    semi_to_finished.set_index('Index', inplace = True)
    
    # Load parameters    
    parameter = pd.read_excel("inputs.xlsx", sheet_name = "Parameters", usecols = 'A:C')
    parameter.set_index('var',inplace = True)
    for var_name in parameter.index:
        locals()[var_name] = parameter.loc[var_name,'value']
        
    CF_ele_table =  pd.read_excel("inputs.xlsx", sheet_name = "Parameters", usecols = 'F:I')
    CF_NREL_table = pd.read_excel("inputs.xlsx", sheet_name = "Parameters", usecols = 'L:O')
    CF_NREL_table.columns = CF_ele_table.columns
    CF_ele_table.set_index('year', inplace = True)    
    CF_NREL_table.set_index('year', inplace = True)
    
    # nrows is always thenumber of items, excluding the heading
    yield_limit_table = pd.read_excel("inputs.xlsx", sheet_name = "Parameters", usecols = 'R:S', nrows = 12)
    yield_limit_table.set_index('sector', inplace = True)
  
    outflow_matrix = generate_outflow_matrix(process,flow)
    net_flow_matrix = generate_net_flow_matrix(process,flow)
    
    ### Yongxian's DMFA scrap results
    df_yx = pd.read_excel("inputs.xlsx", sheet_name = 'Scrap',usecols = 'A,J:M', 
                          nrows = 32, skiprows = 0)
    df_yx.set_index('Year', inplace = True)
    
    scrap_supply_old = pd.read_excel("inputs.xlsx", sheet_name = 'Scrap',usecols = 'A:I', 
                          nrows = 32, skiprows = 82)
    scrap_supply_old.set_index('Year', inplace = True)
    scrap_supply_old = scrap_supply_old.T
    scrap_supply_old.index.name = 'idx'
    
    
    
#%%       
class Process:
    def __init__(self, system, idx):
        self.idx = idx
        self.system = system        
       
        self.name = g.process.loc[idx,"Process"]
        self.stage = g.process.loc[idx,"Stage"]
        #self.CF = g.process.loc[idx,'CF']
        
        self.CF_Import = g.process.loc[idx,'CF_Import']
        self.CF_Import_0 = self.CF_Import
        
        self.inflow = 0
        self.outflow = 0
                   
        self.kgCO2_per_MJ_H2 = self.system.kgCO2_per_MJ_H2
        self.price_H2 = self.system.price_H2
        
        self.inflow_table = pd.DataFrame(columns = ["source","amount","type","fraction"])
        self.inflow_table.set_index("source", inplace = True)
        # Sources have empty inflow table
        
        self.process_track = pd.DataFrame(columns = ["year","target",
                                                     "inflow_0","outflow","inflow_1",
                                                     "import","export"])
        self.process_track.set_index("year", inplace = True) 
        
      
        
    def create_2020_flows(self):
        # Create outflow_table structure for this process
        # Completely based on g.flow_0
        # Essentially serves the operate() function before year 0
        
        self.outflow_table = self.system.flow_0[self.system.flow_0["source"] == self.idx]
        self.outflow_table = self.outflow_table[["target","amount","type"]]
        self.outflow_table.set_index("target", inplace = True)
        self.outflow_table_0 = self.outflow_table.copy()
        self.outflow_table_0['fraction'] = self.outflow_table_0['amount']/self.outflow_table_0['amount'].sum()
                
        if self.idx not in ['Use','W']:
            self.outflow = self.system.consumption_target_0.loc[self.idx,'Target']
        else:
            self.outflow = 0

        if self.idx == "EoL":
            self.inflow = self.outflow - g.EoL_table.loc['import','amount']
            
        self.propagate()

        
    def operate(self):
        '''
        Executed after 2021, including the following phases:
        
        Preparation
            Reset process variables   
             
        Part I: Import inputs and export outputs        
            self.import_amt = self.target - self.inflow
            self.export_amt = max(self.inflow + self.import_amt - self.dom_demand, 0)
            
        Part II Allocate Outflows
            self.allocate()     
            
        Part III Propagate: Generate inflows for the downstream processes
            self.propagate()
        
        '''
        # Preparation
        self.t = self.system.t
        
        ### Global Decarbonization
        if self.stage != 'Fabrication':
            if self.idx in ['R1','R2','C1','C2']:
                self.CF_Import = self.CF_Import_0*(1-self.global_decarb_rate_primary*self.t)
            else:
                self.CF_Import = self.CF_Import_0*(1-self.global_decarb_rate*self.t)
                    
        else:
            self.inflow_table['import_CF'] = self.inflow_table['import_CF'] - \
                self.system.semi_CF['global_decarb_rate']*self.inflow_table['import_CF_0']
            
            self.import_gtg = self.import_gtg_0 * (1 - self.global_decarb_rate*self.t)
        
        if self.stage not in ['Source']:
            self.inflow = self.inflow_table["amount"].sum()
            self.inflow_table["fraction"] = self.inflow_table["amount"]/self.inflow
                
        self.inflow_0 = self.inflow
        self.import_amt = 0
        self.export_amt = 0
        
        
        # Part I: Import inputs and Export outputs
        ### Bayer process: import alumina
        if self.idx == 'M':
            self.import_amt = self.target - self.inflow
            self.export_amt = self.system.Sink.loc['M','Offset']
                    
        
        ### Electrolysis: Import Ore
        elif self.idx == 'E':
            self.import_amt = self.target - self.inflow_0
            
        ### EoL: export scrap
        ### This can be interpreted as the unrecyclable part?
        elif self.idx == 'EoL':            
            R1 = self.system.R1
            R2 = self.system.R2
            R1.inflow_table_init = R1.inflow_table.copy()
            R2.inflow_table_init = R2.inflow_table.copy()
            
            R1.internal_last = R1.internal_scrap_intake
            R2.internal_last = R2.internal_scrap_intake
            self.import_wrought_scrap = 0
            self.import_foundry_scrap = 0
            
            #self.wrought_EoL = min(R1.EoL_scrap_ratio*R1.target, self.inflow)
            #self.foundry_EoL = min(R2.EoL_scrap_ratio*R2.target, self.inflow - self.wrought_EoL)
            
            ### The amount of EoL recycled may be hindered by the amont of available internal scrap
            ### Raise the multiplier to internal_last if Option 1 fares badly in the long run
            #R1.presorted_scrap_target = max(R1.internal_last, min(R1.presorted_scrap_needed, R1.internal_last + 1500))# + 1800
            R1.presorted_scrap_target = min(R1.presorted_scrap_needed, R1.internal_last + self.system.wrought_scrap_import_limit)  
            R2.presorted_scrap_target = min(R2.presorted_scrap_needed, R2.internal_last + self.system.foundry_scrap_import_limit)
            
            ### post-consumer scrap for R1 and R2
            ### If (in rare cases) there are more pre-consumer than targeted amount
            ### The exceeding amount will replace post-consumer scrap
            self.wrought_EoL = R1.presorted_scrap_target*R1.EoL_scrap_ratio/R1.presorted_scrap_ratio - max(R1.internal_last - R1.presorted_scrap_target, 0)
            self.foundry_EoL = R2.presorted_scrap_target*R2.EoL_scrap_ratio/R2.presorted_scrap_ratio - max(R2.internal_last - R2.presorted_scrap_target, 0)            
            
            self.outflow_table.loc['R1','EoL'] = max(self.wrought_EoL,0)
            self.outflow_table.loc['R2','EoL'] = max(self.foundry_EoL,0)            
            
            
            self.import_wrought_scrap = max(R1.presorted_scrap_target - R1.internal_last, 0)                    
            self.import_foundry_scrap = max(R2.presorted_scrap_target - R2.internal_last, 0)               
            
            self.outflow_table.loc['R1','import'] = self.import_wrought_scrap
            self.outflow_table.loc['R2','import'] = self.import_foundry_scrap
            
            self.import_amt = self.import_wrought_scrap + self.import_foundry_scrap
            
            
            self.To_R1 = self.outflow_table.loc['R1','EoL'] + self.import_wrought_scrap
            self.To_R2 = self.outflow_table.loc['R2','EoL'] + self.import_foundry_scrap
            
            self.export_amt = self.inflow+self.import_amt - self.To_R1 - self.To_R2
            
        
        ### Remelting (R1) and Refining (R2)
        ### If pre-consumer scrap is sufficient
        ### Then: self.internal_last + self.external_scrap+self.sweet_required = self.target
        elif self.idx in ['R1','R2']:
            self.dom_primary_received = self.inflow_table.loc['E','amount']                       
                        
            ### Import metal as sweetener
            self.import_metal = 0
                        
            if self.system.push_scrap == False:
                ### Default: If scrap is being pulled, only process the available scrap                
                self.sweet_required = (self.internal_last + self.external_scrap)*(1/self.scrap_ratio - 1)            
            else:
                ### Alternative: If scrap is being pushed, fulfill the target
                self.sweet_required = self.target - (self.internal_last + self.external_scrap)
            
            self.import_amt = max(self.sweet_required - self.dom_primary_received, 0)
            
            self.import_metal = self.import_amt
            
        ### Ingot Casting
        elif self.idx in ['C1','C2']:
            ### Import primary metal
            
            self.import_amt = min(max(self.target - self.inflow,0),
                                  self.target*0.03)
            

            ### Export alloy output            
            self.dom_demand = self.outflow_table['target'].sum()
            self.exp_demand = self.system.Sink.loc[self.idx,'Offset']
            
            self.export_amt = max(self.inflow + self.import_amt - self.dom_demand, 0)
        
        ### Forming
        ### Ensure producing enough semi-products for domestic fabrication
        elif self.stage in ['Forming']:
            
            self.dom_demand = self.outflow_table[self.outflow_table['type']!='Recycle']['demand'].sum()
            self.exp_demand = self.system.Sink.loc[self.idx,'Semi']
            self.value_demand = self.dom_demand + self.exp_demand
            
            ### import casting alloy
            self.capacity_0 = self.inflow*self.efficiency
            if self.capacity_0 > self.value_demand:
                self.import_amt = 0 
            else:
                self.import_amt = (self.value_demand - self.capacity_0)/self.efficiency
        

            self.capacity = (self.inflow+self.import_amt)*self.efficiency #produced by consuming all inflow            
            if 'F' in self.idx: ### deformation
                self.scrap = (self.inflow+self.import_amt) - self.capacity            
            else: ### shape casting
                self.loss = (self.inflow+self.import_amt) - self.capacity
                self.internal = 4*self.loss
                                
            self.export_amt = max(self.capacity - self.dom_demand, 0) #export semi-finished products            
            
        
        ### Fabrication: import semi-components:
        elif self.stage == 'Fabrication':            
            for idx in self.inflow_table.index:
                ### import semi to fill the inflow-target gap                
                self.inflow_table.loc[idx, 'import_amt'] = max(self.inflow_table.loc[idx, 'semi_target'] - self.inflow_table.loc[idx, 'amount'],0)                
                        
            self.inflow_table['import_CO2'] = self.inflow_table['import_amt']*self.inflow_table['import_CF']
            self.import_amt = self.inflow_table['import_amt'].sum()
        
        
        ### Part II Allocate Outflows
        self.allocate()
        
        ### Part III Propagate: Generate inflows for the downstream processes
        self.propagate()
        
        ### Part IV Finalize outflow  
        if self.stage == 'Sink':
            self.outflow = self.inflow
        else:
            self.outflow = self.outflow_table["amount"].sum() # export excluded if process is not sink
        
        self.outflow_table["fraction"] = self.outflow_table["amount"]/self.outflow
                
        ### Stats of Year Beginning             
        self.process_track.loc[self.t, 'target'] = self.target
        self.process_track.loc[self.t, 'inflow_0'] = self.inflow_0 # for R1 and R2, the inflow will change at year end 
        self.process_track.loc[self.t, 'outflow'] = self.outflow
        self.process_track.loc[self.t, 'import'] = self.import_amt
        self.process_track.loc[self.t, 'export'] = self.export_amt
        self.process_track.loc[self.t,'clean_grid_ratio'] = self.clean_grid_ratio
        self.process_track.loc[self.t,'CF'] = self.CF

        
    def allocate(self):                                          
        if self.idx == 'EoL':                            
            self.outflow_table.loc['R1','amount'] = self.To_R1
            self.outflow_table.loc['R2','amount'] = self.To_R2
            self.system.fraction_matrix.loc['EoL','R1'] = self.To_R1/(self.inflow+self.import_amt-self.export_amt)
            self.system.fraction_matrix.loc['EoL','R2'] = self.To_R2/(self.inflow+self.import_amt-self.export_amt)
                
                        
        elif self.idx == 'M':
            self.outflow_table.loc['E', "amount"] = self.target*self.efficiency - self.export_amt
            self.outflow_table.loc['W', "amount"] = self.target*(1 - self.efficiency)
        
        elif self.idx == 'E':            
            for idx in self.outflow_table.index:
                if idx!= 'W':
                    # pull value flows
                    self.system.fraction_matrix.loc[self.idx,idx] = self.outflow_table.loc[idx,'target']/(self.inflow+self.import_amt)
                produced = (self.inflow_0+self.import_amt-self.export_amt)*self.system.fraction_matrix.loc[self.idx,idx]                
                self.outflow_table.loc[idx, "amount"] = produced
       
        ### Allocate Others           
        elif self.stage in ['Forming']:
            for idx in self.outflow_table.index:
                if self.outflow_table.loc[idx,'type']!= 'Recycle':
                    produced = self.outflow_table.loc[idx,'demand']
                if idx == 'R1': #R1 and R2
                    produced = self.scrap*self.R1_ratio
                if idx == 'R2': #R1 and R2
                    produced = self.scrap*self.R2_ratio
                if idx == 'W':
                    produced = self.loss
                
                self.outflow_table.loc[idx, "amount"] = produced
                self.system.fraction_matrix.loc[self.idx,idx] = produced/(self.inflow_0+self.import_amt-self.export_amt)
        
        elif self.stage in ['Casting']:            
            if self.inflow+self.import_amt>self.dom_demand:
                ### Pull             
                for idx in self.outflow_table.index:
                    target = self.outflow_table.loc[idx,'target']                                    
                    produced = target                    
                    self.outflow_table.loc[idx, "amount"] = produced                    
                    self.system.fraction_matrix.loc[self.idx,idx] = target/(self.inflow_0+self.import_amt-self.export_amt)
            else:
                ### Push
                for idx in self.outflow_table.index:
                    target = self.outflow_table.loc[idx,'target']
                    #update fraction
                    self.system.fraction_matrix.loc[self.idx,idx] = target/self.dom_demand
                    produced = (self.inflow_0+self.import_amt)*self.system.fraction_matrix.loc[self.idx,idx]                
                    self.outflow_table.loc[idx, "amount"] = produced
              
        else:
            ### Push
            for idx in self.outflow_table.index:                
                produced = (self.inflow_0+self.import_amt-self.export_amt)*self.system.fraction_matrix.loc[self.idx,idx]                
                self.outflow_table.loc[idx, "amount"] = produced

        
    def propagate(self):
        for target_idx in self.outflow_table.index:
            target = self.system.process_dict[target_idx]                     
            target.inflow_table.loc[self.idx,"amount"] = self.outflow_table.loc[target_idx, "amount"]
            target.inflow_table.loc[self.idx,"type"] = self.outflow_table.loc[target_idx, "type"]
            


    '''-----------------Process Properties-------------------------''' 
    @property
    def internal_scrap(self):
        return self.inflow_table[self.inflow_table['type'].isin(['Recycle'])]['amount'].sum()
   
    @property
    def internal_scrap_intake(self):
        return self.inflow_table_init[self.inflow_table_init['type'].isin(['Recycle'])]['amount'].sum()
   
    @property
    def internal_scrap_generated(self):
        return self.internal_scrap
        
    @property 
    def external_scrap(self):
        return self.inflow_table.loc['EoL','amount']
        
   
    @property
    def presorted_scrap_needed(self):
        return self.target*self.presorted_scrap_ratio
    
    @property
    def EoL_needed(self):
        return self.target*self.EoL_scrap_ratio
   
    @property
    def scrap_intake(self):
        return self.inflow_table_init[self.inflow_table_init.index!='E']['amount'].sum()             
    
   
    @property
    def scrap_per_in(self):
        return self.inflow_table_init[self.inflow_table_init.index != 'E']['amount'].sum()/self.outflow
    
    @property
    def sweet_per_in(self):
        return 1 - self.scrap_per_in
    
    @property
    def wrought_ratio(self):
        return self.wrought_EoL/self.inflow
    
    @property
    def foundry_ratio(self):
        return self.foundry_EoL/self.inflow
    
    @property
    def waste_per_value(self):
        if self.idx in ['M','E','R1','R2','SC1','SC2','SC3']:
            return self.outflow_table.loc['W', 'amount']/self.system.value_target_last
        else:
            return 0
   
    
    @property
    def CF(self):
        #first find kgCO2e per MJ-delivered
        CF_kgCO2_per_MJ = self.NG_ratio*g.kgCO2_per_MJ_NG\
                        + self.ele_ratio*(1-self.clean_grid_ratio)*self.system.kgCO2_per_MJ_ele_dirty\
                        + self.ele_ratio*self.clean_grid_ratio*g.kgCO2_per_MJ_ele_clean\
                        + self.H2_ratio*self.kgCO2_per_MJ_H2\
                        
                        
        CF = CF_kgCO2_per_MJ*self.energy_intensity+ self.CF_process
        return CF
    
    @property
    def CF_heat(self):
        return (self.NG_ratio*g.kgCO2_per_MJ_NG + self.H2_ratio*self.kgCO2_per_MJ_H2)*self.energy_intensity
    
    @property
    def CF_ele(self):
        CF_ele = (self.ele_ratio*(1-self.clean_grid_ratio)*self.system.kgCO2_per_MJ_ele_dirty\
                   + self.ele_ratio*self.clean_grid_ratio*g.kgCO2_per_MJ_ele_clean)\
            *self.energy_intensity
        return CF_ele
    
    @property
    def CF_out(self):
        # CF_kgCO2_per_kg_out
        return self.CF/self.efficiency
    
    @property
    def CO2_heat(self):
        return self.CF_heat*(self.inflow_0+self.import_amt)/1000
                
    @property
    def CO2_ele(self):        
        return self.CF_ele*(self.inflow_0+self.import_amt)/1000
        
    @property
    def CO2_proc(self):
        return self.CF_process*(self.inflow_0+self.import_amt)/1000
        
            
    @property
    def CO2_oprt(self):
        #processed amount is the inflow amount at year beginning
        #self.CO2_oprt = self.CO2_dom_to_dom + self.CO2_dom_to_export
        CO2_oprt = (self.inflow_0+self.import_amt)*self.CF # = (self.outflow+self.export_amt)*self.CF_out
        
        return CO2_oprt
    
    @property
    def CO2_dom_to_dom(self):
        if 'P' in self.idx:
            return self.system.product_table.loc[self.idx,'dom_to_dom_CO2']/self.efficiency
        else:        
            return self.CO2_oprt*self.dom_to_dom_valuable/self.valuable
    
    @property
    def CO2_dom_to_export(self):
        if 'P' in self.idx:
            return self.system.product_table.loc[self.idx,'export_CO2']/self.efficiency
        else:        
            return self.CO2_oprt*self.export_amt/self.valuable
        
    @property
    def CO2_import(self):
        
        CO2_import  = self.import_amt*self.CF_Import
        if self.stage == 'Fabrication':
            CO2_import = self.inflow_table['import_CO2'].sum()
        if self.idx == 'Use':
            CO2_import = self.system.CO2_impt_use*1000
        
        return CO2_import
    
    @property 
    def dom_to_dom_finished(self):
        return self.system.product_table.loc[self.idx, 'dom_to_dom']
        
    @property
    def export_finished(self):
        return self.system.product_table.loc[self.idx, 'export']
    
    @property
    def CO2_import_to_dom(self):
        if 'P' in self.idx:           
            return self.CO2_import*self.dom_to_dom_finished/(self.dom_to_dom_finished + self.export_finished)
        else:
            return self.CO2_import*self.dom_to_dom_valuable/self.valuable
    
    @property
    def CO2_import_to_export(self):
        if 'P' in self.idx:
            return self.CO2_import*self.export_finished/(self.dom_to_dom_finished + self.export_finished)
        else:
            return self.CO2_import*self.export_amt/self.valuable
    
    @property
    def CO2_export(self):
        return self.CO2_dom_to_export +  self.CO2_import_to_export
        
    
    @property
    def CO2(self):
        return self.CO2_oprt + self.CO2_import
    
    @property
    def energy(self):
        # energy intensity: MJ/kg, 
        # outflow: kt
        # energy: MJ
        return self.energy_intensity*10**6*(self.inflow_0+self.import_amt) 
    
    @property
    def cost_per_MJ(self):
        # cost_per_MJ: $/MJ
        return g.price_NG*self.NG_ratio\
            +self.price_H2*self.H2_ratio\
            +g.price_ele_dirty*self.ele_ratio*(1-self.clean_grid_ratio)\
            +g.price_ele_clean*self.ele_ratio*self.clean_grid_ratio
        
    @property
    def cost_per_kg_out(self):
        # energy intensity: MJ/kg-in
        # cost_per_MJ: $/MJ
        return self.energy_intensity*self.cost_per_MJ/self.efficiency
    
    
    @property 
    def cost(self):
        #energy: in MJ
        #unit energy cost: $ per MJ
        #Cost: in Billion $
        cost = self.energy*self.cost_per_MJ/10**9
        # The same as self.valuable*self.cost_per_kg_out/10**3
        return cost
    
    @property
    def efficiency_true(self):
        return (self.inflow+self.import_amt-self.loss-self.scrap)/(self.inflow+self.import_amt+self.internal)
        # (SC1.inflow+SC1.internal)*SC1.efficiency_true = SC1.outflow + SC1.export_amt - SC1.loss

    @property
    def total_outflow(self):
        return self.outflow_table['amount'].sum()+self.export_amt
    
    @property
    def valuable(self):
        
        return self.total_outflow - self.outflow_table[self.outflow_table['type'].isin(['Recycle','Waste'])]['amount'].sum()
    
    @property
    def dom_to_dom_valuable(self):
        return self.outflow -  self.outflow_table[self.outflow_table['type'].isin(['Recycle','Waste'])]['amount'].sum()
    
    @property
    def scrap_amt(self):       
        R1_scrap = self.outflow_table.loc['R1', 'amount'] if 'R1' in self.outflow_table.index else 0
        R2_scrap = self.outflow_table.loc['R2', 'amount'] if 'R2' in self.outflow_table.index else 0
        return R1_scrap + R2_scrap

    @property
    def ctg(self):
        #return the embodied CO2 of the valuable output of a process
        C1 = self.system.C1
        C2 = self.system.C2
        
        E_C1 = C1.inflow_table.loc['E','amount']
        E_C2 = C2.inflow_table.loc['E','amount']
        
        process_dict = self.system.process_dict
        process = self.system.process
        ctg = pd.DataFrame(columns = ['source'])
        ctg.set_index('source',inplace = True)
        
        
        ctg.loc['Operation','type'] = process.loc[self.idx,'Output']
        ctg.loc['Operation','kg/kg-out'] = 1
        ctg.loc['Operation','kgCO2e/kg'] = self.CF_out
        ctg.loc['Operation','kgCO2e/kg-out'] = self.CO2_oprt/self.valuable
        ctg.loc['Operation','$/mt'] = self.cost_per_kg_out*1000
        ctg.loc['Operation','$/mt-out'] = ctg.loc['Operation','kg/kg-out']*ctg.loc['Operation','$/mt']
        
        ctg.loc['Operation','stage'] = self.stage        
        
        
        if self.import_amt > 0:
            import_idx = "Import"
            ctg.loc[import_idx,'type'] = import_idx
            ctg.loc[import_idx,'kg/kg-out'] = self.import_amt/self.valuable
            ctg.loc[import_idx,'kgCO2e/kg'] = self.CO2_import/self.import_amt
            ctg.loc[import_idx,'kgCO2e/kg-out'] = self.CO2_import/self.valuable
            ctg.loc[import_idx,'$/mt'] = 0
            ctg.loc[import_idx,'$/mt-out'] = ctg.loc[import_idx,'kg/kg-out']*ctg.loc[import_idx,'$/mt']
            ctg.loc[import_idx,'stage'] = import_idx

            
        if self.idx in ['R1','R2']:
            for idx in ['E','EoL']:
                source = process_dict[idx]
                ctg.loc[idx,'type'] = self.inflow_table.loc[idx,'type']
                ctg.loc[idx,'kg/kg-out'] = self.inflow_table.loc[idx,'amount']/self.valuable
                ctg.loc[idx,'kgCO2e/kg'] = source.CO2_ctg
                ctg.loc[idx,'kgCO2e/kg-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'kgCO2e/kg']
                ctg.loc[idx,'$/mt'] = source.cost_ctg
                ctg.loc[idx,'$/mt-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'$/mt-out']
                ctg.loc[idx,'stage'] = source.stage
            
            idx = 'internal'
            ctg.loc[idx,'type'] = 'Scrap'
            ctg.loc[idx,'kg/kg-out'] = self.internal_scrap_intake/self.valuable
            ctg.loc[idx,'kgCO2e/kg'] = 0
            ctg.loc[idx,'kgCO2e/kg-out'] = 0
            ctg.loc[idx,'$/mt'] = 0
            ctg.loc[idx,'$/mt-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'$/mt']
            ctg.loc[idx,'stage'] = 'Internal'
            
            idx = 'Import_scrap'
            ctg.loc[idx,'type'] = 'Scrap'
            kg_per_kg_out = (
                self.system.EoL.import_wrought_scrap/self.valuable
                if self.idx == 'R1'
                else self.system.EoL.import_foundry_scrap/self.valuable
                )
            
            ctg.loc[idx,'kg/kg-out'] = kg_per_kg_out
            ctg.loc[idx,'kgCO2e/kg'] = self.system.EoL.CF_Import
            ctg.loc[idx,'kgCO2e/kg-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'kgCO2e/kg']
            ctg.loc[idx,'$/mt'] = 0
            ctg.loc[idx,'$/mt-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'$/mt']
            ctg.loc[idx,'stage'] = 'Source'
                
                    
        #if self.stage in ['Casting','Forming','Fabrication']:
        else:
            for idx in self.inflow_table.index:
                
                source = process_dict[idx]
                ctg.loc[idx,'type'] = self.inflow_table.loc[idx,'type']
                ctg.loc[idx,'kg/kg-out'] = self.inflow_table.loc[idx,'amount']/self.valuable                               
                ctg.loc[idx,'stage'] = source.stage
                
                if self.stage == 'Fabrication':                    
                    if idx == 'F1':                                               
                        if self.idx in ['P2','P3']:
                            C1.inflow_table.loc['E','amount'] = C1.valuable*0.283*E_C1/(E_C1+C1.import_amt)
                                                     
                        else:
                            C1.inflow_table.loc['E','amount'] = 0
                                                                                
                    if idx == 'F2':
                        if self.idx in ['P2','P3']:                            
                            C1.inflow_table.loc['E','amount'] = C1.valuable*0.107*E_C1/(E_C1+C1.import_amt)
                                                  
                        else:
                            C1.inflow_table.loc['E','amount'] = C1.valuable*0.231*E_C1/(E_C1+C1.import_amt)
                            
                    C1.inflow_table.loc['R1','amount'] = C1.inflow -  C1.inflow_table.loc['E','amount']                      
                    ctg.loc[idx,'kgCO2e/kg'] = source.CO2_ctg 
                    ctg.loc[idx,'$/mt'] = source.cost_ctg 
                    C1.inflow_table = C1.inflow_table_original.copy()
                    
                elif 'SC' in self.idx:
                    if idx == 'C2':
                        if self.idx == 'SC1':
                            C2.inflow_table.loc['E','amount'] = C2.valuable*0.15*E_C2/(E_C2+C2.import_amt)
                                                        
                        else:
                            C2.inflow_table.loc['E','amount'] = C2.valuable*0.10*E_C2/(E_C2+C2.import_amt)
                            
                    C2.inflow_table.loc['R2','amount'] = C2.inflow - C2.inflow_table.loc['E','amount']
                    ctg.loc[idx,'kgCO2e/kg'] = source.CO2_ctg
                    ctg.loc[idx,'$/mt'] = source.cost_ctg 
                    C2.inflow_table = C2.inflow_table_original.copy()
                        
                                       
                else:        
                    ctg.loc[idx,'kgCO2e/kg'] = source.CO2_ctg
                    ctg.loc[idx,'$/mt'] = source.cost_ctg 
                    
                ctg.loc[idx,'$/mt-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'$/mt']
                ctg.loc[idx,'kgCO2e/kg-out'] = ctg.loc[idx,'kg/kg-out']*ctg.loc[idx,'kgCO2e/kg']    
                
        ctg.loc[ctg['type'] == 'Castings','stage'] = 'Shape Casting'
        ctg.loc[ctg['stage'] == 'Forming','stage'] = 'Deformation'
                
                    
        return ctg
        
        
    @property
    def CO2_ctg(self):
        #Cradle-to-Gate: Sum of embodied kgCO2e/kg output for each process output
        
        if self.idx in self.system.ctg_memo and self.idx not in ['F1', 'F2', 'SC1','C1','C2']:            
        
            ### Dynamic programming: if already calculated
            ### and does not need to be manually tuned based on condition
            return self.system.ctg_memo[self.idx]
        elif self.idx == 'EoL': #Excluding imported scrap
            return self.CF
        else:
            ### If haven't been calculated or needs to be manually tuned based on condition
            return self.ctg['kgCO2e/kg-out'].sum()

    
    @property
    def cost_ctg(self):
        #Cradle-to-Gate: Sum of embodied $/kg output for each process output
        return self.ctg['$/mt-out'].sum()


#%%                    

class Simulation:
    '''
    To create an instance of the system object:
        sim = Simulation(warmup_period = warmup_period, sim_period = sim_period)
        
        sim.initialize()
        
        sim.run()

    Parameters
    ----------
    warmup_period : scalar(int)
        Default 4: 2020 - 2024
        Grid decarbonization applied to this period through self.evolve()
        global decarbonization applied to this period through process.operate()

    sim_period : scalar(int)
        Default 26: 2025 - 2050
        Decarbonization pathways applied to this period

    Attributes
    ----------
    A: 22*22 dataframe
        Input-Output Matrix
    Sink: 22*5 dataframe
        Consumption matrix
        5 Columns: Use, Semi, Offset, W, Sink
    
    Workflow
    ----------
    __init__()
        Pass on Parameters and create system variables
    initialize()
        Create process objects and generate 2020 flows
    
    
    '''
    
    def __init__(self, warmup_period = 0, sim_period = 0, 
                 yield_update = False, yield_increasing_rate = 0.01,
                 EoU_update = False, recycling_increasing_rate = 0.01,
                 grid_update = False,
                 hydrogen = False, blue = False, PTC = False,
                 electrification = False,
                 technology_update = False, adoption_rate = 0.0385,
                 alternative_trade = False, trade_mode = 'Island', #['Island','Surplus','Deficit']
                 CV = 0, lightweight = False, initial_year = 2020,
                 SA = False, dom_alloy_ratio = 0.85, dom_semi_ratio = 0.9,
                 global_decarb_mode = 'Constant', #[Constant, Frozen, CA2020],
                 push_scrap = False,                  
                 wrought_scrap_import_limit = 1450, #700
                 foundry_scrap_import_limit = 500, #500
                 record_process = False, display = False, stepwise = False, breakdown_emissions = False,
                 optimization = False, optimization_all = False):
        '''Passing on Parameters and create system variables. 
        
        Does not generate flows '''
                
        self.warmup_period = warmup_period
        self.sim_period = sim_period
        self.yield_update = yield_update
        self.EoU_update = EoU_update
        self.grid_update = grid_update
        self.hydrogen = hydrogen
        self.blue = blue
        self.PTC = PTC
        self.electrification = electrification    
        self.technology_update = technology_update 
        self.alternative_trade = alternative_trade
        if alternative_trade == True:
            self.trade_mode = trade_mode
        
        self.record_process = record_process
        self.display = display
        self.stepwise = stepwise
        self.breakdown_emissions = breakdown_emissions
        
        self.yield_increasing_rate = yield_increasing_rate
        self.recycling_increasing_rate = recycling_increasing_rate
        self.adoption_rate = adoption_rate
        
        self.CV = CV
        self.lambda_rate = g.lambda_rate
        self.weight = 1
        self.lightweight = lightweight
        self.initial_year = initial_year
        self.last_year = initial_year + warmup_period + sim_period                
                
        self.year = self.initial_year
        self.t = 0
        self.env = simpy.Environment()
        self.process = g.process.copy()
        self.process.drop(columns = ['CF Ref'], inplace = True)
        self.A = g.A.copy()
        
        self.global_decarb_mode = global_decarb_mode
        self.push_scrap = push_scrap
        self.semi_CF = g.semi_CF
        
        self.flow_0 = g.flow.copy()
        self.Sink = g.Sink.copy()
        self.Sink_unit = g.Sink.copy() #Final demand in unit (1 unit = 1 kg in 2020)        
        self.Sink_0 = g.Sink.copy()
        self.Sink_unit_0 = g.Sink.copy()
        self.consumption_target = g.consumption_target.copy()
        self.consumption_target_0 = g.consumption_target.copy()
        
        
        self.SA = SA
        self.dom_alloy_ratio = dom_alloy_ratio
        self.dom_semi_ratio = dom_semi_ratio
        
        self.optimization = optimization
        self.optimization_all = optimization_all
        
        
        
        # End_use table and post-consumer scrap settings
        self.end_use = pd.read_excel("inputs.xlsx", sheet_name = "end_use",usecols= 'A:G', nrows = 8)
        self.end_use.set_index("idx", inplace = True)
        self.recyclable_percent = 1
                
        # EoL at t = 0
        self.EoU = 0
        self.new_post_0 = 0
        self.old_post_0 = g.df_yx.loc[2020,'old_post_0']
        self.available_EoL = g.df_yx.loc[2020,'Total']
        self.collecting_upper_limit = 0.95
        
        # Pre-consumer setting
        self.wrought_scrap_import_limit = wrought_scrap_import_limit
        self.foundry_scrap_import_limit = foundry_scrap_import_limit
        
        
        # Create Essential tables
        self.process_list = []
        self.process_dict = {}
                
        self.system_track = pd.DataFrame(columns = ["now","t","year"])
        self.system_track.set_index("now",inplace = True)
        
        self.system_deets = self.system_track.copy()
       
        ### Energy settings
        self.kgCO2_per_MJ_ele_dirty = g.CF_ele_table.loc[2020, 'CF_dirty']/1000
        
        self.kgCO2_per_MJ_H2 = g.kgCO2_per_MJ_H2_green if self.blue == False else g.kgCO2_per_MJ_H2_blue   
        self.system_clean_grid_ratio = g.CF_ele_table.loc[2020, 'clean_grid_ratio']         
        
        if self.blue == False:
            self.price_H2 = g.price_H2_green if self.PTC == False else g.price_H2_green_PTC
        else:
            self.price_H2 = g.price_H2_blue if self.PTC == False else g.price_H2_blue_PTC 
       
        ### Supply Chain settings 
        self.export_scale = 1.0
        self.export_scale_0 = self.export_scale
        
  
    
    def initialize(self):
        '''
        (1) Creating Process Objects and assign their initial attributes        
        (2) Generating 2020 flows based on g.flow        
        (3) Create inflow_table for all processes based on 2020 flows        
        (4) Define Waste Matrix based on 2020 flows
        '''

            
        # (1) Creating process objects and assign initial attributes
        for idx in self.process.index:
            p = Process(system = self, idx = idx)
            self.process_list.append(p)
            self.process_dict[idx] = p
            
            p.internal = 0
            p.scrap = 0
            p.loss = 0
            
            ### Energy and emission attributes
            p.energy_intensity = g.process.loc[idx,'MJ_per_kg']#Delivered energy per kg-in
            p.energy_intensity_0 = p.energy_intensity
            p.NG_ratio = g.process.loc[idx,'NG_ratio']
            p.NG_ratio_0 = p.NG_ratio
            p.ele_ratio = g.process.loc[idx,'ele_ratio']
            p.ele_ratio_0 = p.ele_ratio
            p.H2_ratio = 0
            p.clean_grid_ratio = g.process.loc[idx,'clean_grid_ratio']
            p.clean_grid_ratio_0 = p.clean_grid_ratio
            
            p.CF_process = g.process.loc[idx,'CF_process']
            p.CF_process_0 = p.CF_process
            if p.idx == 'E':
                p.PFC_CF_process = g.PFC_CF_process
                p.PFC_CF_process_0 = p.PFC_CF_process
                p.CO2_CF_process = g.CO2_CF_process
                p.CO2_CF_process_0 =  p.CO2_CF_process
                p.inert_ratio = 0
            p.CF_0 = p.CF    
            
            ### import gate to gate assumed to be equal to dom
            p.import_gtg = p.CF # kgCO2e/kg-in
            # will be converted to kgCO2e/kg-out in product_table
            p.import_gtg_0 = p.import_gtg
                        
            ### Consumption demand growth rate for finished products
            if p.stage == 'Fabrication':
                p.lambda_rate = g.product_trade_table.loc[p.idx,'lambda']
                
            ### Global decarbonization rate
            if self.global_decarb_mode == 'Constant':
                ### Calculated from global_emissions.py
                p.global_decarb_rate = 0.0140  # decreasing by 0.76 pct point/year
                self.semi_CF['global_decarb_rate'] = 0.0140
                p.global_decarb_rate_primary = 0.0081 # decreasing by 0.81 pct point/year
                
            elif self.global_decarb_mode == 'Frozen':
                p.global_decarb_rate = 0
                self.semi_CF['global_decarb_rate'] = 0.000
            elif self.global_decarb_mode == 'CA2020':
                if p.CF_Import!=0:
                    p.global_decarb_rate = (p.CF_Import - p.CF_Import*0.711)/p.CF_Import/30
                else:
                    p.global_decarb_rate = 0.000
                self.semi_CF['global_decarb_rate'] = (g.semi_CF['CF_ctg'] - g.semi_CF['CF_CA'])/g.semi_CF['CF_ctg']/30
            
            
        # (2) Generating 2020 Flows
        for p in self.process_list:
            p.create_2020_flows() # instead of p.operate()
                                    
        self.fraction_matrix = self.generate_fraction_matrix() #has only been executed once, based on initial outflow
        self.fraction_matrix_0 = self.fraction_matrix.copy()
        self.value_target = self.Sink['Use'].sum()+self.Sink['Semi'].sum()+self.Sink['Offset'].sum()
        
        # (3) Define inflow_table_0 (after R1 and R2 receive recycle flows) 
        for p in self.process_list:                        
            p.inflow_table["fraction"] = p.inflow_table["amount"]/p.inflow_table["amount"].sum()
            if p.idx in self.consumption_target.index:
                p.dom_ratio = p.inflow_table['amount'].sum()/self.consumption_target_0.loc[p.idx,'Target']
            
            if p.idx in ['R1','R2']:
                p.inflow_table_init = p.inflow_table
                
                p.internal_scrap_0 = p.internal_scrap
                p.external_scrap_0 = p.external_scrap
                p.scrap_per_in_0 = p.scrap_per_in
                p.scrap_ratio = getattr(g, p.idx+"_scrap_ratio")
                p.scrap_ratio_0 = p.scrap_ratio
                p.EoL_scrap_ratio = getattr(g, p.idx+"_EoL_scrap_ratio")
                p.EoL_scrap_ratio_0 = p.EoL_scrap_ratio
                p.presorted_scrap_ratio = p.scrap_ratio - p. EoL_scrap_ratio
                p.presorted_scrap_ratio_0 = p.presorted_scrap_ratio
                
                inflow_received = self.consumption_target_0.loc[p.idx,'Target']                
                scrap_received = p.internal_scrap + p.external_scrap
                sweet_received = inflow_received - scrap_received
                dom_sweet_received = p.inflow_table.loc['E','amount']
                p.dom_sweet_ratio = dom_sweet_received/sweet_received
                p.dom_sweet_ratio_0 = p.dom_sweet_ratio
             
               
            if p.stage == 'Fabrication':
                for source_idx in p.inflow_table.index:
                    p.inflow_table.loc[source_idx,'efficiency'] = g.fab_efficiencies.loc[source_idx,p.idx]
                    p.inflow_table.loc[source_idx, 'import_CF'] = self.semi_CF.loc[source_idx,'CF_ctg']
                    p.inflow_table.loc[source_idx, 'import_CF_0'] = self.semi_CF.loc[source_idx,'CF_ctg']

                p.dom_to_dom_ratio = g.product_trade_table.loc[p.idx, 'dom_to_dom_ratio']
                p.dom_to_dom_ratio_0 = p.dom_to_dom_ratio
                p.dom_to_dom_ratio_delta = 0.0
                p.export_ratio = g.product_trade_table.loc[p.idx, 'export_ratio']
                p.export_ratio_0 = p.export_ratio
                p.export_ratio_delta = 0.0
                #p.import_ratio = g.product_trade_table.loc[p.idx, 'import_ratio']
                #p.import_ratio_0 = p.import_ratio
                
                p.dom_con_target = g.product_trade_table.loc[p.idx, 'dom_consumed']
                p.dom_con_target_0 = p.dom_con_target
                p.export_target = g.product_trade_table.loc[p.idx, 'export']
                p.export_target_0 = p.export_target
                p.production_target = p.dom_con_target*p.dom_to_dom_ratio + p.export_target
                p.production_target_0 = p.production_target
                
                
                                  
            
            if p.stage in ['Forming','Fabrication']:
                if self.fraction_matrix_0.loc[p.idx,'R1']+self.fraction_matrix_0.loc[p.idx,'R2']>0:        
                    p.R1_ratio = self.fraction_matrix_0.loc[p.idx,'R1']/(self.fraction_matrix_0.loc[p.idx,'R1']+self.fraction_matrix_0.loc[p.idx,'R2'])
                    p.R2_ratio = self.fraction_matrix_0.loc[p.idx,'R2']/(self.fraction_matrix_0.loc[p.idx,'R1']+self.fraction_matrix_0.loc[p.idx,'R2'])
                else:
                    p.R1_ratio = 0
                    p.R2_ratio = 0
                    
            if p.stage != 'Sink':
                total_outflow = (p.outflow_table_0['amount'].sum()+self.Sink.loc[p.idx,'Semi']+self.Sink.loc[p.idx,'Offset'])
                p.efficiency = 1-p.outflow_table_0[p.outflow_table_0['type'].isin(['Recycle','Waste'])]['amount'].sum()/total_outflow
                p.efficiency_0 = p.efficiency
                
            else:
                p.efficiency = 1
                
            if p.idx == 'EoL':                
                p.scrap_import_rate = g.EoL_table.loc['import','amount']/(self.R1.outflow - self.R1.internal_scrap - p.outflow_table.loc['R1','amount'] + g.EoL_table.loc['import','amount'])
            
            
            
            
            p.inflow_table_0 = p.inflow_table.copy()
            
                    
        # (4) Define Waste matrix
        self.Gamma = pd.DataFrame(0.0, index=self.A.index, columns=self.A.columns)
        for idx in ['M','E','R1','R2','SC1','SC2','SC3']:
            self.Gamma.loc[idx,idx] = 1 - self.process_dict[idx].efficiency
            
        
        # Initial Conditional settings
        ### Trade Scenarios
        if self.alternative_trade == True:
            '''          
            3 #Alternative Scenarios; by 2050
            (1) Island:
                - domestic supply reliance increase to 100%
                - export targets drop to 0
            
            (2) Surplus (Mercantile): 
                - domestic supply reliance increase to 100%
                - export targets increased by 100% across the supply chain compared to basecase 2050
            
            (3) Deficit (Post-industry):
                - domestic supply reliance drop to half of 2020 level
                - export targets drop to 0
                
            Two key variable types:
            sim.export_scale: 
                - a factor multiplied on the export targets of all flows except for finished-products
                - Default 1.0
            p.export_ratio:
                - a factor multiplied on the domestic consumption of finished-products
                - default value varies by product category
                
            
            '''
            
            
            if self.trade_mode in ['Island','Deficit']:
                ### reach 0 before 2050
                export_scale_delta = - 0.04                 
            elif self.trade_mode == 'Surplus':
                ### increased by 100% before 2050
                export_scale_delta = 0.04
            else:
                export_scale_delta = 0.0
                
            self.export_scale_delta = export_scale_delta
       
                        
            if self.trade_mode in ('Island','Surplus'): 
                ### R1 and R2 have 100% domestic sweetener by 2050
                self.R1.dom_sweet_ratio_delta = (1.0 - self.R1.dom_sweet_ratio_0)/26
                self.R2.dom_sweet_ratio_delta = (1.0 - self.R2.dom_sweet_ratio_0)/26
             
                
            if self.trade_mode == 'Deficit':
                ### R1 and R2's domestic sweetener relianceis drop to half of the 2020 level by 2050.
                self.R1.dom_sweet_ratio_delta = - 0.5*self.R1.dom_sweet_ratio_0/26
                self.R2.dom_sweet_ratio_delta = - 0.5*self.R2.dom_sweet_ratio_0/26

            for col in self.A.columns:
                p = self.process_dict[col] 
                p.dom_ratio_delta = 0.0
                                
                if col not in ['M','EoL','R1','R2']:                
                    dom_ratio_0 = self.A[col].sum()
                    
                    if self.trade_mode in ['Surplus','Island']:   
                        ### Increase to 100% domestic reliance by 2050
                        p.dom_ratio_delta = (1/dom_ratio_0 - 1)/26
                    elif self.trade_mode == 'Deficit':
                        ### reduce to 50% of domestic reliance of 2020
                        p.dom_ratio_delta = - 0.5/26
                        

                ### End_Use                                                        
                if 'P' in col:
                                     
                    if self.trade_mode in ['Island', 'Surplus']:
                        ### These ratios domestic reliance Use, not P1 - P8!      
                        if self.trade_mode in ['Surplus','Island']:   
                            dom_to_dom_ratio_delta = (1 - p.dom_to_dom_ratio_0)/26
                                                
                        if self.trade_mode == 'Island':
                            ### export reduce to 0 by 2050
                            export_ratio_delta = - p.export_ratio_0/26
                        if self.trade_mode == 'Surplus':
                            ### export increase by 100% by 2050
                            export_ratio_delta = p.export_ratio_0/26
                            
                            
                    if self.trade_mode == 'Deficit':
                        ### reduce to 50% of domestic reliance of 2020
                        dom_to_dom_ratio_delta = - 0.5*p.dom_to_dom_ratio_0/26
                        ### export reduce to 0% by 2050
                        export_ratio_delta = - p.export_ratio_0/26
                        
                    
                    p.dom_to_dom_ratio_delta = dom_to_dom_ratio_delta
                    p.export_ratio_delta = export_ratio_delta
                    
                ### Extra settings
                self.wrought_scrap_import_limit = np.inf
                self.foundry_scrap_import_limit = np.inf
                        
            
        ### Yield improvement settings                
        if self.yield_update == True:
            for idx in g.yield_limit_table.index:
                p = self.process_dict[idx]
                p.yield_limit = g.yield_limit_table.loc[p.idx, 'limit']
                p.yield_increasing_rate  = (p.yield_limit - p.efficiency_0)/26
                
        if self.EoU_update == True:
            self.R1.EoL_scrap_ratio_1 = g.R1_EoL_scrap_ratio_1
            # See Input.xlsx - Calculator - [C4]
            # 0.4 in first submission
            # The above target is obtained by assuming  0.457197936 of PCRC in 2050
            self.EoL_increasing_rate = (self.R1.EoL_scrap_ratio_1 - self.R1.EoL_scrap_ratio_0)/26
        
        ### Sensitivity analysis settings
        if self.SA == True:
            self.alternative_trade = False        
            ### Alloy
            for col in ['F'+str(i) for i in range(1,5)]:                                      
                self.A.loc['C1',col] = self.dom_alloy_ratio
            for col in ['SC'+str(i) for i in range(1,4)]:
                self.A.loc['C2',col] = self.dom_alloy_ratio
               
            ### Semi
            for col in ['P'+str(i) for i in range(1,9)]:
                self.A[col] = self.A[col]*(self.dom_semi_ratio)/g.A[col].sum()
                
            ### Update the default 2020 consumption target based on the overwritten A   
            self.consumption_target = generate_consumption_target_init(self.A, self.Sink)
            for i in range(1,9):
                self.consumption_target.loc['P'+str(i)] = g.consumption_target.loc['P'+str(i)]
        
        ### Scrap optimization settings
        if self.optimization == True:
            self.alloy_demand_dict = {} 
            self.alloy_supply_dict = {}
            self.forming_scrap_dict = {}
            self.fab_scrap_dict = {}
            
            
            for semi in ['Sheet','Extrusion','Forgings','Castings']:
                self.alloy_demand_dict[semi] = pd.DataFrame().rename_axis('Alloy')
                self.alloy_supply_dict[semi] = pd.DataFrame().rename_axis('Alloy')
                self.forming_scrap_dict[semi] = pd.DataFrame().rename_axis('Alloy')
                self.fab_scrap_dict[semi] = pd.DataFrame().rename_axis('Alloy')
                
            for col in list(range(2020,2051)):
                self.end_use[col] = 0.0
                # Record available EoL scrap at the beginning of a year
                # Each year summation should be equal to system_track[new_post]
                
   
            self.alloy_comp_dict = {}
            index_list = ['P'+str(i) for i in range(1,9)]
            
            for idx in index_list:
                df = pd.read_excel('alloys.xlsx', sheet_name=idx, index_col = 0, nrows = 12)
                #df.index = [f"{df.index.name} {row}" for row in df.index]                            
                self.alloy_comp_dict[idx] = df
       
            
    def iterate(self):
        # Before this iteration, the EoU at the **end** of **this** year already generated (scheduled)
        # which does not depend on what happens this year.
        
        ### To generate consumption_target (domestic demand), need to:            
            ### - update A matrix based on current flow srructure, which is characterized by outflow_matrix
            ### - update self.Sink based on new producr dewmand
            ### - recalculate consumption_target based on updated Leontief and Sink
        
        if self.env.now>0:            
            
            self.evolve() #grid transition and decarbonization pathways inside
            self.update_A() # Update A matrix before anything else!
            
            if self.push_scrap == True and self.t >= self.warmup_period + 1: 
                self.push_post_scrap()
                
            
            
            self.generate_input()
            
        
        
        if self.optimization == True:
            if self.env.now == 0:
                self.alloy_composition_matrix = self.calculate_alloy_composition()
            
            self.scrap_supply_to_excel()
            if self.optimization_all == True:
                self.optimize() 
            elif self.year in [2020]: #[2020, 2025, 2035, 2050]:
                self.optimize()
                           
        self.set_targets()
        
                    
        for p in self.process_list:
            p.operate()
        
        # Update the inflow_table of R1 and R2 at the year end
        for idx in ['R1','R2']:
            p = self.process_dict[idx]                                             
            p.inflow = p.inflow_table["amount"].sum()
            p.inflow_table["fraction"] = p.inflow_table["amount"]/p.inflow
        
        self.C1.inflow_table_original = self.C1.inflow_table.copy()
        self.C2.inflow_table_original = self.C2.inflow_table.copy()
        
        
        self.schedule_EoU()
        
        self.record_system_status()
        if self.record_process == True:        
            self.record_process_status()
        if self.stepwise == True:
            self.record_detailed_status()
            
                    
        
        self.year = self.year+1
        self.t = self.t + 1
                
        ### The scheduled events will happen right now
    
    def evolve(self):
        ### Dynamic Behavior        
        ### System-wise (domestic) grid transition 
        if self.year<=2050:
            self.system_clean_grid_ratio = g.CF_ele_table.loc[self.year, 'clean_grid_ratio']
            self.kgCO2_per_MJ_ele_dirty = g.CF_ele_table.loc[self.year, 'CF_dirty']/1000
        else:
            self.kgCO2_per_MJ_ele_dirty = max(0, self.kgCO2_per_MJ_ele_dirty - 2.0/1000)
            self.system_clean_grid_ratio = min(self.clean_grid_ratio +0.002, 1.0)
            
        for p in self.process_list: 
            if p.idx != 'E':
                p.clean_grid_ratio = self.system_clean_grid_ratio
                
        self.E.clean_grid_ratio = min(1.0, self.E.clean_grid_ratio+ 1.06/100)
        
                
            
        if self.lightweight == True:
            self.weight = self.weight - 0.01
        
            
        
        ### Implement strategies 
        #if self.grid_update == True:
            #if self.env.now>= 2023:
                #self.improve_grid()
        
        if self.env.now>=self.warmup_period+1:
            if self.yield_update == True:
                self.improve_yield()
                                
            if self.EoU_update == True:
                self.improve_EoL()
                                
            if self.grid_update == True:
                self.improve_grid()
                
            if self.hydrogen == True:
                self.transition_hydrogen()
            
            if self.electrification == True:
                self.electrify()
                
            if self.technology_update == True:
                self.adopt_technology()

            if self.alternative_trade == True:
                self.execute_alternative_trade()
                

                
                
                    
        ### point feeder
        ### y = E.PFC_CF_process, x = E.inert_ratio
        ### y = y_0*(1-x)*e^{-t} with even better control technology
        #self.E.PFC_CF_process = self.E.PFC_CF_process_0*(1-self.E.inert_ratio)*np.exp(-self.t*0.03)
        self.E.PFC_CF_process = self.E.PFC_CF_process_0*(1-self.E.inert_ratio)
        self.E.CO2_CF_process = (1 - self.E.inert_ratio)*self.E.CO2_CF_process_0
        self.E.CO2_CF_process = max(self.E.CO2_CF_process,0)                
        self.E.CF_process = self.E.PFC_CF_process + self.E.CO2_CF_process
        #print("System updated in year ", self.year)
        
    
    def generate_input(self):
        '''
        Determine overall consumption target for each process
        
        Set up exogeneous inflows:
            Mining
                By default set to 0
            EoL
                new post-consumer scrap
                    self.new_post_0 = self.EoU_last                      
                old post-consumer scrap
                    Load YX's scrap data\n                    
                    Linearly decreasing after 2050
        '''
        
        # Finished-Product Demand
        for i in range(1,9):
            p = self.process_dict['P'+str(i)]
                        
            ### Fluctuating around trend
            dom_con_target = p.dom_con_target_0 * (1+p.lambda_rate*self.t*np.random.normal(1,self.CV))                        
            p.dom_con_target = max(dom_con_target, 0)
            
            p.dom_to_dom_target = p.dom_con_target*p.dom_to_dom_ratio
            p.export_target = p.dom_con_target * p.export_ratio
            p.production_target = p.dom_to_dom_target + p.export_target 
            
            self.Sink_unit.loc['P'+str(i), 'Use'] = p.production_target 
        
        # Other Demands
        self.delta_Semi = self.Sink_0['Semi']*self.lambda_rate*np.random.normal(1,self.CV)*self.t
        self.delta_Offset = self.Sink_0['Offset']*self.lambda_rate*np.random.normal(1,self.CV)*self.t
        
        self.Sink_unit['Semi'] = (self.Sink_unit_0['Semi'] + self.delta_Semi).clip(lower=0)*self.export_scale
        self.Sink_unit['Offset'] = (self.Sink_unit_0['Offset'] + self.delta_Offset).clip(lower=0)*self.export_scale
        
                
        # Calculating consumption target                       
        self.Sink = self.Sink_unit*self.weight
        self.consumption_target, self.Sink = generate_consumption_target(self.A, self.Gamma,self.Sink)
        
        # Virgin Production
        self.process_dict['M'].inflow = 0                     
        
        # Post Consumer Scrap Demand
        self.required_EoL = self.consumption_target.loc['EoL','Target']
        
        ### Available new scrap generated after simulation starts
        self.new_post_0 = self.EoU_last
        
        ### Available old scrap from YX's data
        if self.year<2050:
            self.old_post_0 = g.df_yx.loc[self.year, 'old_post_0']
        else:
            self.old_post_0 = max(self.old_post_0 - 50,0)
        
        self.available_EoL = g.df_yx.loc[self.year, 'Total']
        self.process_dict['EoL'].inflow = self.EoU_last + self.old_post_0
        
        

    
    def set_targets(self):    
        ### Determine process #targets and Adjusting Propagating fractions
        R1 = self.R1
        R2 = self.R2
        
        for p in self.process_list:  
            if p.idx not in ['Use','W']:
                p.target = self.consumption_target.loc[p.idx,'Target']
            elif p.idx == 'Use':
                p.target = self.product_target
            else:
                p.target = 0               
            
                               
        for p in self.process_list: 
            if p.idx == 'E':
                                       
                E_R1 = R1.target*self.A.loc['E','R1']                                       
                E_R2 = R2.target*self.A.loc['E','R2']  
                E_C1 = self.consumption_target.loc['C1', 'Target']*self.A.loc['E','C1']
                E_C2 = self.consumption_target.loc['C2', 'Target']*self.A.loc['E','C2']
                  
                p.outflow_table.loc['R1','target'] = E_R1
                p.outflow_table.loc['R2','target'] = E_R2
                p.outflow_table.loc['C1','target'] = E_C1
                p.outflow_table.loc['C2','target'] = E_C2
                p.outflow_table.loc['W','target'] = 0
                p.target = p.outflow_table['target'].sum()/p.efficiency
                
                M_E = p.target*self.A.loc['M','E']
                M_export = self.Sink.loc['M','Offset']
                self.process_dict['M'].target = (M_E+M_export)/self.process_dict['M'].efficiency
                
            
            elif p.stage in ['Casting']:
                for target_idx in p.outflow_table.index:
                    p.outflow_table.loc[target_idx,'target'] = self.process_dict[target_idx].target*self.A.loc[p.idx,target_idx]
            
            elif p.stage in ['Forming']:
                for target_idx in p.outflow_table.index:
                    if target_idx != 'W':
                        p.outflow_table.loc[target_idx,'demand'] = self.process_dict[target_idx].target*self.A.loc[p.idx,target_idx]
                    else:
                        p.outflow_table.loc[target_idx,'demand'] = 0
            
            elif p.stage in ['Fabrication']:
                p.fab_target = self.Sink.loc[p.idx,'Use']
                p.inflow_table['semi_target'] = p.target * p.inflow_table_0['fraction']
                #for source_idx in p.inflow_table.index:
                    #p.inflow_table.loc[source_idx,'semi_target'] = p.target*p.inflow_table_0.loc[source_idx,'fraction']
                    ##p.inflow_table.loc[source_idx,'semi_target'] = p.fab_target/p.efficiency*p.inflow_table_0.loc[source_idx,'fraction']

            elif p.idx == 'Use':
                p.inflow_table['target'] = self.Sink.loc[p.inflow_table.index, 'Use']
                #for source_idx in p.inflow_table.index:
                    #target = p.system.Sink.loc[source_idx,'Use'] #domestic_production_target
                    #p.inflow_table.loc[source_idx,"target"] = target

    def schedule_EoU(self):
        '''
        Generate post-consumer scrap
        
        For the same year
            Directly add to self.EoL_same_year
        For the following years
            Execute self.generate_EoU(current_year = self.year)
        
        All EoU will be generated **after** this iteration, but **before** the next iteration
        '''
                
        ### (1) Consumables generated at the beginning of this year will expire at the year end
        container_available = self.product_table.loc['P7','consumed']
        container_recycled = container_available*self.end_use.loc['P7', 'recyc_rate']
        self.EoL_same_year = container_recycled
        if self.optimization == True:
            if self.t + 1 <= self.warmup_period+self.sim_period:  
                #self.end_use.loc['P7',self.year + 1] += container_recycled
                self.end_use.loc['P7',self.year + 1] += container_available
        
        for idx in g.production_idx_list:
            mean_life = self.end_use.loc[idx,"life"]
            std_life = self.end_use.loc[idx,"std"] 
            fraction_1 = generate_probability(1, mean_life = mean_life, std_life = std_life)
            available_amt = self.product_table.loc[idx,'consumed']*fraction_1
            recycled_amt = available_amt*0.95
            
            self.EoL_same_year += recycled_amt
            
            if self.optimization == True: 
                if self.t + 1 <= self.warmup_period + self.sim_period:                
                    #self.end_use.loc[idx, self.year + 1] +=  recycled_amt
                    self.end_use.loc[idx, self.year + 1] +=  available_amt
            
        
        ### (2) EoU of other categories will be generated at the **End** of the **next** year
        ### the scheduled updates will happen at the end of this iteration           
        self.env.process(self.generate_EoU(current_year = self.year))
                
        ### (3) The EoU amount at this year-end
        ### Note that self.EoU has already been updated before this iteration
        ### will be fed into the EoL node at the beginning of next year
        ### In other words, system_track.loc[n, 'EoL'] <=> system_track.loc[n-1,'EoU']
        self.EoU_last = self.EoU+self.EoL_same_year
        
        
        ### (4) Before the scheduled events happen, first reset the remaining EoU        
        self.EoU = 0
    
    def generate_EoU(self,current_year):
        ### Update the EoU for the **Next** Year
        ### Immediately generate this year's amount, then yield
                
        production_table = self.product_table.copy()
        
        for i in range(1, self.last_year - current_year+1): 
            # for the remainder of the simulation
            ### Schedule the amount to be added to self.EoU_scheduled
            ### First batch is scheduled to be added at the end of next year
            ### and enter the system at the beginning of the following year
            
            for idx in g.production_idx_list:
                new_product = production_table.loc[idx, "consumed"]
        
                collect_rate = self.end_use.loc[idx,"collect_rate"]
                recyclable = self.end_use.loc[idx,"recyclable"]
                recycling_rate = collect_rate*recyclable
                new_recycling_rate = recycling_rate
                mean_life = self.end_use.loc[idx,"life"]
                std_life = self.end_use.loc[idx,"std"]                                
                
                                
                ### update recycle rate
                #expire_year = current_year+i
                #implementation_year = self.initial_year + self.warmup_period+1
                #if self.EoU_update == True and expire_year>implementation_year:
                    #if current_year < implementation_year: # if produced before implementation
                        #collect_rate = min(collect_rate+(expire_year - implementation_year)*self.recycling_increasing_rate,self.collecting_upper_limit)                        
                    #else: # if produced after implementation
                        #collect_rate = min(collect_rate+(expire_year - current_year)*self.recycling_increasing_rate,self.collecting_upper_limit)
                    #new_recycling_rate = collect_rate*recyclable
                    
                                
                ### Schedule the amount to be added to self.EoU
                fraction_i = generate_probability(i+1, mean_life = mean_life, std_life = std_life)
                available_amt = fraction_i*new_product
                recycled_amt = available_amt*new_recycling_rate
                self.EoU = self.EoU + recycled_amt
    
                
                if self.optimization == True:
                    if self.t+1 <= self.warmup_period + self.sim_period:
                        ### Scheduled at the beginning of the next year
                        #self.end_use.loc[idx, self.year+1] +=  recycled_amt
                        self.end_use.loc[idx, self.year+1] +=  available_amt
                
                #scheduled_year = current_year + i
                #print("%.3f"%amount, ' of ', idx, " from ", current_year, " added to ", self.year, '[{}]'.format(scheduled_year))
                    
            yield(self.env.timeout(1))

    
    def record_process_status(self):        
        ### Stats at Year-End

        for p in self.process_list: 
            p.inflow_1 = p.inflow 
            p.process_track.loc[self.env.now,"inflow_1"] = p.inflow_1
            
            self.process.loc[p.idx, 'Target'] = p.target
            self.process.loc[p.idx, 'Inflow_0'] = p.inflow_0
            self.process.loc[p.idx, 'Import'] = p.import_amt
            self.process.loc[p.idx, 'Export'] = p.export_amt
            self.process.loc[p.idx, 'Outflow'] = p.outflow
            self.process.loc[p.idx, 'Inflow_1'] = p.inflow_1
            self.process.loc[p.idx, 'clean_grid_ratio'] = p.clean_grid_ratio
            self.process.loc[p.idx, 'CO2_oprt'] = p.CO2_oprt/1000 #Convert from kt to Mt
            self.process.loc[p.idx, 'CO2_impt'] = p.CO2_import/1000 #Convert from kt to Mt
            self.process.loc[p.idx, 'CO2'] = p.CO2/1000 #Convert from kt to Mt
                        
            
            
            if self.electrification == True or self.hydrogen ==  True:
                self.process.loc[p.idx,'NG_ratio'] = p.NG_ratio
                self.process.loc[p.idx,'ele_ratio'] = p.ele_ratio
                
        W = self.process_dict['W']
        for source_idx in W.inflow_table.index:
            self.process.loc[source_idx, 'waste_per_value'] = W.inflow_table.loc[source_idx,'amount']/self.value_target
                
                
        ### (2) System-Level    
    def record_system_status(self):
        ### Stats at Year-End
        self.system_track.loc[self.env.now,"t"] = self.t
        self.system_track.loc[self.env.now,"year"] = self.year
        
        self.system_track.loc[self.env.now, "product_target"] = self.product_target                       
        #self.system_track.loc[self.env.now, "semi_target"] = self.semi_product_target
        #self.system_track.loc[self.env.now, "target"] = self.product_target + self.semi_product_target
        
        #self.system_track.loc[self.env.now, "product_produced"] = self.product_fabricated 
        #self.system_track.loc[self.env.now, "semi_exported"] = self.semi_product_export
        
        
        #self.system_track.loc[self.env.now, "total_output"] = self.total_output
        self.system_track.loc[self.env.now, "consumed"] = self.domestic_consumed_product    
        #self.system_track.loc[self.env.now, "net_output"] = self.product_fabricated + self.semi_product_export
        
        self.system_track.loc[self.env.now,"primary"] = self.primary_production
        self.system_track.loc[self.env.now,"secondary"] = self.secondary_production
        
        #self.system_track.loc[self.env.now,"pre_available"] = self.internal_scrap_intake
        #self.system_track.loc[self.env.now,"pre_import"] = self.EoL.import_amt
        #self.system_track.loc[self.env.now,"pre_recycled"] = self.internal_scrap_intake+self.EoL.import_amt
        
        ###EoL Scrap Status
        #self.system_track.loc[self.env.now,"EoL_collected"] = self.process_dict['EoL'].inflow
        self.system_track.loc[self.env.now,"EoL"] = self.available_EoL
        self.system_track.loc[self.env.now,"EoL_collected"] = self.process_dict['EoL'].inflow
        self.system_track.loc[self.env.now,"EoL_recycled"] = self.recycled_EoL        
        self.system_track.loc[self.env.now,"old_post_0"] = self.old_post_0
        self.system_track.loc[self.env.now,"new_post"] = self.new_post_0
        self.system_track.loc[self.env.now,"new_post_1"] = self.EoU_last
        
        
        #self.system_track.loc[self.env.now,"external"] = self.total_input
        
        #self.system_track.loc[self.env.now,"import"] = self.total_import
        #self.system_track.loc[self.env.now,"export"] = self.total_export
                
        #self.system_track.loc[self.env.now,"alloy ingot"] = self.wrought_alloy + self.cast_alloy        
        #self.system_track.loc[self.env.now,"waste"] = self.total_waste
        
        #self.system_track.loc[self.env.now,"material_efficiency"] = self.material_efficiency
        
        ### Carbon and Cost Status
        ### [Note] sim.dom_to_dom_CO2 + sim.dom_to_export_CO2 = sim.domestic_CO2
        ###        sim.domestic_CO2 + sim.import_CO2 = sim.total_CO2
        ### sim.dom_to_dom_CO2 + sim.import_to_dom_CO2 + sim.export_CO2 = sim.total_CO2
        self.system_track.loc[self.env.now,"CO2"] = self.total_CO2
        self.system_track.loc[self.env.now,"import_CO2"] = self.import_CO2
        self.system_track.loc[self.env.now,"dom_CO2"] = self.domestic_CO2

        if self.breakdown_emissions == True:                
            self.system_track.loc[self.env.now,'dom_to_dom_CO2'] = self.dom_to_dom_CO2
            self.system_track.loc[self.env.now,'import_to_dom_CO2'] = self.import_to_dom_CO2
            self.system_track.loc[self.env.now,"export_CO2"] = self.export_CO2
        
        self.system_track.loc[self.env.now,"cost"] = self.energy_cost
        
        if self.env.now>0:
            self.system_track.loc[self.env.now,"Sys_inflow"] = self.system_inflow
        
        else:
            self.system_track.loc[self.env.now,"Sys_inflow"] = self.total_input+self.total_import+self.internal_scrap_intake
        self.system_track.loc[self.env.now,"Sys_outflow"] = self.system_outflow    
        
         
        
        ### Display Status
        if self.display == True:
            if self.env.now<=self.warmup_period:
                print("Warmup ",self.env.now, " (", self.year, ") "," finished", sep = '')
            if self.env.now == self.warmup_period:
                print("Warm up complete")
                print()
            if self.env.now>self.warmup_period:
                print("Year ",self.t, " (", self.year, ") "," finished", sep = '')
                
    def record_detailed_status(self):
        '''
        Record detailed system/process variables
        
        Only executed in stepwise mode
        '''
        self.system_deets.loc[self.env.now,"t"] = self.t
        self.system_deets.loc[self.env.now,"year"] = self.year
        
        self.system_deets.loc[self.env.now, "clean_grid_ratio"] = self.system_clean_grid_ratio
        self.system_deets.loc[self.env.now, "gCO2_per_MJ_ele"] = self.kgCO2_per_MJ_ele*1000
        
        self.system_deets.loc[self.env.now, "inert_anode_ratio"] = self.E.inert_ratio
        self.system_deets.loc[self.env.now, "E_epsilon"] = self.E.energy_intensity
        self.system_deets.loc[self.env.now, "E_CF"] = self.E.CF
        
        if self.optimization == True:
            self.system_deets.loc[self.env.now, "alloy_produced"] = self.alloy_produced
            self.system_deets.loc[self.env.now, "avaialable_post"] = self.total_post[self.year].sum()
            self.system_deets.loc[self.env.now,"available_pre"] = self.internal_scrap_intake
       
        
    #Options    
    def improve_yield(self, limit = 0.95):
        for idx in g.yield_limit_table.index:
            p = self.process_dict[idx]
            ### Increase yield efficiency to p.yield_limit
            p.efficiency = p.efficiency + p.yield_increasing_rate
            
            
            if p.stage == 'Fabrication':
                self.fraction_matrix.loc[p.idx,'Use'] = p.efficiency
                self.fraction_matrix.loc[p.idx,'R1'] = (1-p.efficiency)*p.R1_ratio
                self.fraction_matrix.loc[p.idx,'R2'] = (1-p.efficiency)*p.R2_ratio

            
        
    def improve_EoL(self):                
        ###Upcycling
        rate = self.EoL_increasing_rate
        
        self.R1.EoL_scrap_ratio = self.R1.EoL_scrap_ratio + rate
        self.R1.presorted_scrap_ratio = self.R1.presorted_scrap_ratio - rate/2
        self.R1.scrap_ratio = self.R1.EoL_scrap_ratio + self.R1.presorted_scrap_ratio
        
    def push_post_scrap(self):
        ### Set higher production target for remelted metal (produced from R1)
        self.A.loc['E','C1']  = max(self.A.loc['E','C1'] - g.A.loc['E','C1']/26, 0)
        self.A.loc['R1','C1'] = max(g.A['C1'].sum() - self.A.loc['E','C1'],0)
        ### Ensure the system produce the same amount of wrought alloy as of BAU 2050
        self.A.loc['E', 'R1'] = self.A.loc['E', 'R1'] + (0.0981114031261555 - g.A.loc['E', 'R1'])/26
        
        
    
    def improve_grid(self):
        '''
        overwrite system_clean_grid_ratio and carbon intensity of dirty electricity
        '''
        
        self.system_clean_grid_ratio = g.CF_NREL_table.loc[self.year, 'clean_grid_ratio']
        self.kgCO2_per_MJ_ele_dirty = g.CF_NREL_table.loc[self.year, 'CF_dirty']/1000
        for p in self.process_list: 
            if p.idx != 'E':
                p.clean_grid_ratio = self.system_clean_grid_ratio
        self.E.clean_grid_ratio = min(1.0, self.E.clean_grid_ratio+ 2.2/100)          
        
        
    def transition_hydrogen(self):
        #only replace NG        
        for idx in self.process_dict:
            
            p = self.process_dict[idx]
            if p.stage in ['Source','Processing','Casting','Forming','Fabrication']:
                rate = 1/11*p.NG_ratio_0       
                
                p.H2_ratio = min(p.H2_ratio+rate,1-p.ele_ratio)
                p.NG_ratio = 1 - p.H2_ratio - p.ele_ratio
            
            
    def electrify(self):
        #replace NG with current grid         
        for idx in self.process_dict:            
            p = self.process_dict[idx]
            if p.stage in ['Source','Processing','Casting','Forming','Fabrication']:
                rate = 1/11*p.NG_ratio_0       
                p.ele_ratio = min(p.ele_ratio+rate,1 - p.H2_ratio)               
                p.NG_ratio = 1 - p.H2_ratio - p.ele_ratio
                
    def adopt_technology(self):
        self.E.inert_ratio = self.E.inert_ratio + self.adoption_rate
        self.E.inert_ratio = min(self.E.inert_ratio,1)
        self.E.energy_intensity = self.E.energy_intensity_0*(1 - self.E.inert_ratio)\
                                + self.E.energy_intensity_0*(1+9.48/100)*self.E.inert_ratio
        
    
    def execute_alternative_trade(self):
        
        def clip(x, lb = 0.0, ub = 1.0):
            return min(max(x, lb), ub)
        
        ### A-matrix and sector-level settings
        for col in self.A.columns:
            p = self.process_dict[col]
            
            p.dom_ratio_multiplier_t = 1 + p.dom_ratio_delta*(self.t - self.warmup_period)
            
            self.A[col] = g.A[col]*p.dom_ratio_multiplier_t
            
            ### End-Use Only
            if 'P' in col:                
                p.dom_to_dom_ratio = p.dom_to_dom_ratio_0 + p.dom_to_dom_ratio_delta*(self.t - self.warmup_period)
                p.dom_to_dom_ratio = clip(p.dom_to_dom_ratio)
                p.export_ratio = p.export_ratio_0 + p.export_ratio_delta*(self.t - self.warmup_period)
                p.export_ratio = clip(p.export_ratio)
                
        ### system-level settings
        self.export_scale = clip(self.export_scale + self.export_scale_delta, ub = 2.0)        
        self.R1.dom_sweet_ratio = clip(self.R1.dom_sweet_ratio + self.R1.dom_sweet_ratio_delta)
        self.R2.dom_sweet_ratio = clip(self.R2.dom_sweet_ratio + self.R2.dom_sweet_ratio_delta)
        
        self.A = self.A.clip(upper = 1.0)
        
        

    # ____________ Core Simulation Mechanisms_____________________                            
    def clock(self):
        while True:
            self.iterate()
            yield self.env.timeout(1)        
        
    def run(self):
        '''
        Run the simulation until the target end time\n        
        No detailed tracking
        
        Execute the simulation stepwise to track detailed stats
        
        '''
        if self.stepwise == False:
            self.env.process(self.clock())
            self.env.run(until = self.warmup_period + self.sim_period+1)
            print("Complete {:.2f} {:.2f} {:.2f}".format(self.total_CO2, self.system_inflow, self.system_outflow),
                  abs(self.system_inflow - self.system_outflow)<0.0001)
            
        else:
            self.env.process(self.clock())
            current_time = 0
            end_time = self.warmup_period + self.sim_period + 1
            
            while current_time < end_time:
                self.env.run(until = current_time + 1)
                current_time = self.env.now

                #print(f"Time: {current_time:.2f}")
                
            print("Complete {:.2f} {:.2f} {:.2f}".format(self.total_CO2, self.system_inflow, self.system_outflow),
                  abs(self.system_inflow - self.system_outflow)<0.0001)
        
    def optimize(self):
        self.alloy_demand_dict = self.update_alloy_dict(self.alloy_demand_matrix, self.alloy_demand_dict)
        self.alloy_supply_dict = self.update_alloy_dict(self.alloy_supply_matrix, self.alloy_supply_dict)
        self.forming_scrap_dict = self.update_alloy_dict(self.forming_scrap_matrix, self.forming_scrap_dict)
        self.fab_scrap_dict = self.update_alloy_dict(self.fab_scrap_matrix, self.fab_scrap_dict)
        self.alloy_to_excel(self.alloy_demand_dict, 'domestic_Al_demand.xlsx')
        self.alloy_to_excel(self.alloy_supply_dict, 'domestic_Al_supply.xlsx')
        self.alloy_to_excel(self.forming_scrap_dict, 'forming_scrap.xlsx')
        self.alloy_to_excel(self.fab_scrap_dict, 'fab_scrap.xlsx')
        
        #print(self.alloy_demand_dict['Sheet'].loc['B&C Sheet 3xxx', self.year])
        #print(self.forming_scrap_dict['Sheet'].loc['B&C Sheet 3xxx', self.year])
    
    
                
    def update_alloy_dict(self, df, dictionary):                
        for semi in ['Sheet','Extrusion','Forgings','Castings']:                                          
            for product in df.columns:
                df_alloy_comp = self.alloy_comp_dict[product]
                product_name = df_alloy_comp.index.name
                
                for alloy in df_alloy_comp.index:
                    comp = df_alloy_comp.loc[alloy, semi]
                    mass = df.loc[semi, product]
                    amount = comp*mass #either demand or supply
                    
                    if amount > 0.00000001:
                        label = f"{product_name} {semi} {alloy}"
                        
                        if label == 'Auto Sheet 6xxx':
                            dictionary[semi].loc[f"{label} 6005C", self.year] = amount/2
                            dictionary[semi].loc[f"{label} 6016", self.year] = amount/2
                        else:
                            dictionary[semi].loc[label, self.year] = amount
        return dictionary
    
    def alloy_to_excel(self, dictionary, file_path):

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for sheet_name, df in dictionary.items():
                if df.empty:
                    ref_df = self.alloy_demand_dict[sheet_name]
                    df = pd.DataFrame(0.0, index=ref_df.index, columns=ref_df.columns)
                
                df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=True, header=True)

        self.format_excel(file_path)
        
    def scrap_supply_to_excel(self):
        
        file_path = 'domestic_scrap_supply.xlsx'
        self.scrap_supply.to_excel(file_path, sheet_name = 'Sheet1', index = True)
        
        self.format_excel(file_path)

    def format_excel(self, file_path):
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
               
            # Widen first column
            ws.column_dimensions[get_column_letter(1)].width = 30
    
            # Apply formatting to all index cells (A2 down)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left')
                    cell.border = Border()  # remove all borders
    
        wb.save(file_path)
                


        
    '''-----------------System Properties-------------------------'''            
    def product_intensity(self, metric = 'carbon'):
        if metric == 'carbon':
            column = 'kgCO2e/kg-out'
        else:
            column = '$/mt-out'
        df = pd.DataFrame(columns = ['source','type'])
        df.set_index('source', inplace = True)
        index_list = ['Operation','Import','F1','F2','F3','F4','SC1','SC2','SC3']
        for idx in index_list:
            
            if idx not in ['Operation','Import']:
                if idx in ['SC1','SC2','SC3']:
                    df.loc[idx,'type'] = self.process.loc[idx,'Output']
                else:
                    df.loc[idx,'type'] = self.process.loc[idx,'Display']
            else:
                df.loc[idx,'type'] = idx
        
        df.loc['SC','type'] = 'Shape Castings'
                    
        for i in range(1,9):
            p = self.process_dict['P'+str(i)]
            ctg = p.ctg
            for idx in index_list:
                if idx in ctg.index:
                    df.loc[idx,'P'+str(i)] = ctg.loc[idx,column]
                else:
                    df.loc[idx,'P'+str(i)] = 0
        
        
            df.loc['SC','P'+str(i)] = df[df['type'] == 'Castings']['P'+str(i)].sum()
                            
        
            
        df.drop(df[df['type'] == 'Castings'].index, inplace = True)
        df.rename(columns={'type': 'Category'}, inplace=True)
                
        return df
    
    
    @property
    def EoL(self):
        return self.process_dict['EoL']
    
    @property
    def E(self):
        return self.process_dict['E']    
    
    @property
    def R1(self):
        return self.process_dict['R1']    
    
    @property
    def R2(self):
        return self.process_dict['R2']
    
    @property
    def C1(self):
        return self.process_dict['C1']
    
    @property
    def C2(self):
        return self.process_dict['C2']
    
    @property
    def F1(self):
        return self.process_dict['F1']
    
    @property
    def P1(self):
        return self.process_dict['P1']
        
        
            
               
    @property
    def kgCO2_per_MJ_ele(self):
        return self.kgCO2_per_MJ_ele_dirty*(1-self.system_clean_grid_ratio)\
            +  g.kgCO2_per_MJ_ele_clean*self.system_clean_grid_ratio
        
        
    @property
    def total_CO2(self):
        #return self.process["CO2"].sum()
        total_CO2 = 0
        for p in self.process_list:
            total_CO2 += p.CO2/1000
        #total_CO2 += self.CO2_impt_use
        return total_CO2
    
    @property
    def import_CO2(self):
        return sum(p.CO2_import for p in self.process_list) / 1000
    
    @property
    def domestic_CO2(self):
        #domestic_CO2 = 0
        #for p in self.process_list:
            #domestic_CO2 += p.CO2_oprt/1000
        
        #return domestic_CO2
        return sum(p.CO2_oprt for p in self.process_list) / 1000
    
    @property
    def dom_to_dom_CO2(self):
        non_use = sum(p.CO2_dom_to_dom for p in self.process_list if p.idx not in ['W', 'Use'])/1000
        use = 0
        return non_use +use
    
    @property
    def dom_to_export_CO2(self):
        return sum(p.CO2_dom_to_export for p in self.process_list if p.idx not in ['W', 'Use'])/1000
    
    
    @property
    def import_to_dom_CO2(self):
        non_use = sum(p.CO2_import_to_dom for p in self.process_list if p.idx not in ['W', 'Use'])/1000
        use = self.process_dict['Use'].CO2_import/1000 
        return   non_use + use
    
    @property
    def export_CO2(self):
        return sum(p.CO2_export for p in self.process_list if p.idx not in ['W', 'Use'])/1000
    
    @property
    def primary_production(self):
        return self.E.outflow - self.E.outflow_table.loc['W','amount']
    
    @property
    def secondary_production(self):
        return self.R1.outflow_table.loc['C1','amount'] + self.R2.outflow_table.loc['C2','amount']
    
       
    @property
    def recycled_EoL(self):
        return self.EoL.wrought_EoL + self.EoL.foundry_EoL
    
    @property 
    def primary_consumed(self):
        #Double check: should = self.sweetener_consumed + self.casting_virgin
        domestic_primary = self.E.valuable
        import_primary = (self.R1.import_metal + self.R2.import_metal
                          + self.C1.import_amt + self.C2.import_amt)
        return domestic_primary + import_primary 
    
    @property
    def sweetener_consumed(self):
        domestic_sweet = self.E.outflow_table.loc[self.E.outflow_table.index.isin(['R1', 'R2']), 'amount'].sum()
        import_sweet = self.R1.import_metal + self.R2.import_metal
        return domestic_sweet + import_sweet
    
    @property
    def casting_virgin(self):
        return (self.C1.inflow_table.loc['E','amount']+self.C2.inflow_table.loc['E','amount']
                +self.C1.import_amt + self.C2.import_amt)
    
    @property
    def remelted_metal(self):
        return self.C1.inflow_table.loc['R1','amount']
    
    @property
    def recycled_metal(self):
        return self.C2.inflow_table.loc['R2','amount']
            
    
    
    @property 
    def PCRR(self):
        return self.recycled_EoL/self.available_EoL
    
    @property
    def PCRC(self):        
        return (self.remelted_metal*self.R1.EoL_scrap_ratio + self.recycled_metal*self.R2.EoL_scrap_ratio)\
            /(self.casting_virgin + self.remelted_metal + self.recycled_metal)
    
    @property
    def EoL_recycling_rate(self):
        return self.recycled_EoL/self.process_dict['EoL'].inflow
    
    @property
    def total_recycled_scrap(self):
         return self.flow[self.flow["type"]=="Recycle"]["amount"].sum()
     
    
    @property
    def total_input_target(self):
        return self.consumption_target.loc['M',"Target"] + self.consumption_target.loc['EoL',"Target"]
        
    
    @property
    def source_demand(self):
        return self.consumption_target.loc['M','Target']+self.consumption_target.loc['EoL','Target']
    
    @property
    def product_target(self):
        # deomestic production target of final products
        # Not necessarily the domestic *demand* of final products!
        # it consists of
        ### product fabricated from domestically producwd semi self.Sink['Use'].sum()
        ### product_fabricated from imported semi (see inflow_table of P1-P8)
        return self.Sink['Use'].sum()
        
    @property
    def semi_product_target(self):
        # deomestic production target of semi products
        # Not necessarily the *demand* of semi products!
        return self.Sink['Semi'].sum()
    
    @property
    def semi_product_produced(self):
        return self.process[self.process['Stage'] == 'Fabrication']['Inflow_1'].sum() +self.semi_product_export
        
    @property
    def semi_product_export(self):
        semi_product_export = 0 
        for idx in self.process.index:
            if self.process.loc[idx,"Stage"] == 'Forming':
                semi_product_export += self.process.loc[idx,"Export"]
        return semi_product_export
    
    @property
    def semi_product_import(self):
        semi_product_import = 0
        for idx in self.process.index:
            if self.process.loc[idx,"Stage"] == 'Fabrication':
                semi_product_import += self.process.loc[idx,"Import"]
        return semi_product_import
    
    @property
    def total_input(self):        
        return self.process_dict['M'].inflow + self.process_dict['EoL'].inflow
    
    @property
    def source_import(self):
        return self.process_dict['M'].import_amt + self.process_dict['EoL'].import_amt
    
    @property
    def total_import(self):
        total_import = 0
        for idx in self.process_dict:
            total_import += self.process_dict[idx].import_amt
        for idx in self.product_table.index:
            total_import += self.product_table.loc[idx,'import']
        
        return total_import
        
    @property
    def total_export(self):
        # Should be = self.semi_product_export + self.product_export
        total_export = 0
        for idx in self.process_dict:
            total_export += self.process_dict[idx].export_amt
        for idx in self.product_table.index:
            total_export += self.product_table.loc[idx,'export']
        return total_export
    
    @property
    def product_fabricated(self): # products fabricated domestically
        produced = self.process_dict['Use'].inflow
            
        return produced
    
    @property
    def domestic_consumed_product(self):
        return self.product_table['consumed'].sum()
    
    @property
    def product_import(self):        
        return self.product_table['import'].sum()
        
    
    @property
    def product_export(self):                
        return self.product_table['export'].sum()
        
    @property
    def total_output(self):
        return self.product_fabricated - self.product_export + self.total_export
    
    @property
    def total_value(self):
        total_value = 0 
        for p in self.process_list:
            total_value += p.valuable
        return total_value
            
    @property
    def total_waste(self):
        return self.process_dict['W'].inflow
    
    @property
    def wrought_alloy(self):
        C1 = self.process_dict['C1']
        return C1.outflow+C1.export_amt
    
    @property
    def foundry_alloy(self):
        C2 = self.process_dict['C2']
        return C2.outflow+C2.export_amt

    @property
    def alloy_produced(self):
        return self.wrought_alloy + self.foundry_alloy
    
    @property
    def alloy_intake(self):
        alloy_intake = 0
        for idx in ['F1','F2','F3','F4','SC1','SC2','SC3']:
            alloy_intake += self.process_dict[idx].import_amt + self.process_dict[idx].inflow
        return alloy_intake
            
    @property
    def downstream_efficiency(self):
        net_valuable = self.product_fabricated+self.semi_product_export
        
        return net_valuable/(self.alloy_intake+self.semi_product_import)
    
    @property 
    def net_output(self):
        return self.product_fabricated + self.total_export - self.product_export - self.EoL.export_amt
    
    @property
    def net_input(self):
        return self.system_inflow - self.product_import - self.EoL.export_amt
    
    @property
    def material_efficiency(self):        
        return self.net_output/self.net_input
    
    @property
    def carbon_intensity(self):
        CO2 = self.total_CO2 #Unit: Mt
        #Unit of self.output: kt
        return CO2*1000/self.net_output
    
    
    @property
    def internal_scrap_intake(self):
        R1 = self.process_dict['R1']
        R2 = self.process_dict['R2']
        
        #should be equal to
        #return R1.internal_scrap_intake + R2.internal_scrap_intake
        return R1.inflow_0+R2.inflow_0-R1.inflow_table[R1.inflow_table['type']!='Recycle']['amount'].sum()-R2.inflow_table[R2.inflow_table['type']!='Recycle']['amount'].sum()
    
    @property 
    def internal_scrap_generated(self):
        R1 = self.process_dict['R1']
        R2 = self.process_dict['R2']
        
        #should be equal to
        #return R1.internal_scrap_generated + R2.internal_scrap_generated
        return R1.inflow+R2.inflow-R1.inflow_table[R1.inflow_table['type']!='Recycle']['amount'].sum()-R2.inflow_table[R2.inflow_table['type']!='Recycle']['amount'].sum()
        
    @property
    def system_inflow(self):
        return self.total_input+self.total_import+self.internal_scrap_intake
        
    @property
    def system_inflow_0(self):
        return self.total_input+self.total_import+self.system_track.loc[self.env.now - 2, 'recycled']
    
    @property
    def system_outflow(self):
        return self.domestic_consumed_product+self.internal_scrap_generated+self.total_waste+ self.total_export
    
    @property
    def outflow_matrix(self):
        return generate_outflow_matrix(self.process,self.flow)
    
    @property
    def IO_matrix(self):
        return generate_IO_matrix(self.A, self.Gamma, self.Sink)
    
    @property
    def semi_matrix(self):
        ### dom-to-dom semi-shipments
        df = self.IO_matrix.copy()
        df = df.loc[df.index.to_series().str.startswith(('F', 'SC')), 
                    df.columns.to_series().str.startswith('P')]  
        
        return df
    
    @property
    def forming_scrap_matrix(self):
        df = self.semi_matrix.copy()
        
        for row in df.index:
            # (1) Normlaize each row by row sum
            ### Equivalent to df.div(df.sum(axis=1), axis=0)
            row_sum = df.loc[row].sum()
            df.loc[row] =  df.loc[row]/row_sum
                        
            # (2) Multiply each row by the corresponding p.scrap_amt
            df.loc[row] = df.loc[row]*self.process_dict[row].scrap_amt
            
        df = self.aggregate_alloy(df)
            
        return df
    
    @property
    def fab_scrap_matrix(self):
        df = self.semi_matrix.copy()
        
        for col in df.columns:
            # (1) Normalize each column by col sum
            ### Equivalent to df = df.div(df.sum(axis=0), axis=1)
            col_sum = df[col].sum()
            df[col] = df[col]/col_sum
            
            # (2) Multiply each col by the corresponding p.scrap_amt
            df[col] = df[col]*self.process_dict[col].scrap_amt
            
        df = self.aggregate_alloy(df)
        
        return df
        
        
    @property
    def dom_alloy_demand_matrix(self):
        ### domestic alloys ended up as inflow to P1 - P8
        ### i.e., U.S. produced alloys used for domestic fabrication (including exported finished products)
        df = self.semi_matrix.copy()
        
        for idx in df.index:
            semi_source = self.process_dict[idx]
            df.loc[idx] = df.loc[idx]/semi_source.efficiency
        
        df = self.aggregate_alloy(df)
                
        return df
    
    def aggregate_alloy(self, df):
        # Step 1: Rename 'F1' to 'Sheet'
        df = df.rename(index={'F1': 'Sheet'})
        
        # Step 2: Combine F2 and F3 into 'Extrusion'
        df.loc['Extrusion'] = df.loc['F2'] + df.loc['F3']
        df = df.drop(['F2', 'F3'])
        
        # Step 3: Rename 'F4' to 'Forgings'
        df = df.rename(index={'F4': 'Forgings'})
        
        # Step 4: Combine SC1, SC2, SC3 into 'Castings'
        df.loc['Castings'] = df.loc[['SC1', 'SC2', 'SC3']].sum()
        df = df.drop(['SC1', 'SC2', 'SC3'])
        
        # Optional: Reorder if needed
        df = df.reindex(['Sheet', 'Extrusion', 'Forgings', 'Castings'])
        
        return df
    
    
          
    
    @property
    def alloy_demand_matrix(self):
        ### For each category alloys, semi, and finished products that are produced domestically,
        ### Assuming exports have the same alloy composition as domestic shipments       
        
        df = self.dom_alloy_demand_matrix.copy()
        dom_wrought_total = df.loc[['Sheet', 'Extrusion', 'Forgings']].sum().sum()
        for row in ['Sheet', 'Extrusion', 'Forgings']:
            df.loc[row] = df.loc[row]/dom_wrought_total*self.consumption_target.loc['C1','Target']
            
        dom_foundry_total = df.loc['Castings'].sum()
        df.loc['Castings'] =  df.loc['Castings']/dom_foundry_total*self.consumption_target.loc['C2','Target']
        
        
        return df
    
    
    def calculate_alloy_composition(self):
        df = self.alloy_demand_matrix.copy()
        semi_idx_dict = {'Sheet':'F1',
                         'Extrusion':'F2',
                         'Forgings':'F4',
                         'Castings':'SC1'}
        
        for col in df.columns:            
            for row in df.index:
                semi_idx = semi_idx_dict[row]
                
                
                
                forming_efficiency = self.process_dict[semi_idx].efficiency
                
                if semi_idx in self.process_dict[col].inflow_table.index:
                    fabricating_efficiency = self.process_dict[col].inflow_table.loc[semi_idx,'efficiency']
                else:
                    fabricating_efficiency = 1
                
                df.loc[row, col] = df.loc[row, col]*forming_efficiency*fabricating_efficiency
                      
        
        df_normalized = df.div(df.sum(axis=0), axis=1)
        
        return df_normalized
    
    
   
    @property 
    def total_post(self):
        '''
        Returns
        -------
        df : rows: P1 - P8, columns: years
            Post-consumer scrap for each product category in each year

        '''
        df = g.scrap_supply_old.copy()
        
        for idx in df.index:
            for col in [i for i in range(2020, 2051)]:
                if col <= 2020 + self.warmup_period + self.sim_period:
                    df.loc[idx, col] = df.loc[idx, col] + self.end_use.loc[idx, col]
                else:
                    df.loc[idx, col] = 0
     
        return df
    
    @property
    def scrap_supply(self):
        '''
        Identical to self.total_post with formatted indices
        '''
        
        df = self.total_post.copy()
        df.index = self.end_use['category']
        
        #combined_list = ['Auto', 'Transp.', 'Cons. Dur.']
        
        # # Generate the new label
        # label = ' + '.join([s for s in combined_list])
        
        # # Add the combined row
        # df.loc[label] = df.loc[combined_list].sum() 
        # # Get position of first row to combine
        # pos = df.index.get_loc(combined_list[0])
        # # Drop the original rows
        # df.drop(index=combined_list, inplace = True)
        # # Move new row to the original position
        # df = df.iloc[[*range(pos), -1] + list(range(pos, len(df)-1))]
        
        df.index = df.index + " EOL Scrap"
        df.index.name = 'Sector'
        
        return df
    
    @property
    def alloy_supply_matrix(self):
        '''
        Alloy content in EoL scrap
        '''
        df = pd.DataFrame(columns = [f'P{i}' for i in range(1,9)])   
        
        for col in df.columns:
            available_scrap = self.total_post.loc[col,self.year]
            for row in ['Sheet','Extrusion','Forgings','Castings']:
                alloy_ratio = self.alloy_composition_matrix.loc[row,col]
                alloy_content = available_scrap*alloy_ratio
                df.loc[row,col] = alloy_content
                 
        return df
            
                
                
        
        
        
    @property
    def net_flow_matrix(self):
        return generate_net_flow_matrix(self.process,self.flow)
        
    
    
    @property
    def aggregate_NG_ele_ratio(self):
        NG = 0
        ele = 0
        for idx in self.process_dict:
            p = self.process_dict[idx]
            NG += p.energy*p.NG_ratio/10**6
            ele += p.energy*p.ele_ratio/10**6
            
        return NG/ele
                    

    
    
    def generate_fraction_matrix(self):
        row_sums = self.outflow_matrix.sum(axis = 1)
        row_sums = np.where(row_sums == 0, 1, row_sums)
        fraction_matrix = self.outflow_matrix.div(row_sums, axis =0)
        fraction_matrix.drop(['Use','W'],inplace = True)
        
        
        return fraction_matrix
       

    @property
    def sim_track(self):
        sim_track = self.system_track.copy()
        sim_track['year'] = pd.to_numeric(sim_track['year'])
        sim_track = sim_track[sim_track["t"]>=1]
        
        sim_track.reset_index(inplace = True)
        sim_track.set_index("year", inplace = True)
        return sim_track
    
    @property
    def product_table(self):
        product_table = self.process_dict['Use'].inflow_table.copy()
        
        # the amount of product received by Use from fabrication outflows
        #product_table['received'] = product_table['amount'] # = sum should be = Use.inflow
        product_table.drop(columns = ['amount'], inplace = True)
        
        for idx in product_table.index:
            p = self.process_dict[idx]
            
            # Calculate domestic consumotion by adding import
            received = p.valuable 
            consumed = p.dom_con_target*self.weight
            product_table.loc[idx,'received'] = received                
            product_table.loc[idx,'consumed'] = consumed
                    
            # domestic production ends up in domestic consumption
            # V^dom_i = min(V_i, D_i) = min(V_i, F_i^Use)
            dom_to_dom = min(p.valuable, p.dom_con_target*self.weight*p.dom_to_dom_ratio)
            export_amt = received - dom_to_dom
            product_table.loc[idx,'dom_to_dom'] = dom_to_dom                    
            product_table.loc[idx,'export'] = export_amt
            
            # import
            import_amt = max(consumed  - dom_to_dom,0)
            product_table.loc[idx,'import'] = import_amt
            #product_table.loc[idx,'import'] = max(consumed + export - received,0)
            
            # import ctg
            efficiency = self.process_dict[idx].efficiency_0
            ### Initialize import_ctg as import_gtg, measured in kgCO2e/kg-out
            import_ctg = self.process_dict[idx].import_gtg/efficiency
            
            for semi in g.semi_to_finished.index:
                semi_CF = self.semi_CF.loc[semi,'CF_ctg']
                semi_prop = g.semi_to_finished.loc[semi, idx]
                import_ctg += semi_CF*semi_prop
                
            product_table.loc[idx,'import_ctg'] = import_ctg
            
            # import CO2
            product_table.loc[idx,'import_CO2'] = import_ctg * import_amt
            product_table.loc[idx,'export_CO2'] = self.process_dict[idx].CF*export_amt
            product_table.loc[idx,'dom_to_dom_CO2'] = self.process_dict[idx].CF*dom_to_dom
      
        
        return product_table
          
    @property
    def CO2_impt_use(self):
        return self.product_table['import_CO2'].sum()/1000
    
    @property
    def energy_cost(self):
        energy_cost = 0
        for idx in self.process_dict:
            p = self.process_dict[idx]
            
            energy_cost += p.cost
        
        return energy_cost
    
    @property
    def targets(self):
        targets = pd.DataFrame(columns = ["real","Leontief","equal"])
        for p in self.process_list:
            try:
                targets.loc[p.idx] = [p.target,self.consumption_target.loc[p.idx,"Target"],
                                      (p.target-self.consumption_target.loc[p.idx,"Target"])<0.000001]
            except:
                pass
        return targets
    
    @property
    def scrap_allocation(self):
        R1 = self.R1
        R2 = self.R2
        df = self.EoL.outflow_table[['EoL', 'import']].copy()
        df.loc['R1','presorted needed'] = R1.presorted_scrap_needed
        df.loc['R1','presorted available'] = R1.internal_last
        df.loc['R1','EoL needed'] = R1.EoL_needed
                
        df.loc['R2','presorted needed'] = R2.presorted_scrap_needed
        df.loc['R2','presorted available'] = R2.internal_last
        df.loc['R2','EoL needed'] = R2.EoL_needed
        
        df = df[['EoL needed','EoL','presorted needed','presorted available','import']]
        df = df.apply(lambda x: round(x, 2))
        
        return df
    
    @property
    def process_status_table(self):
        df = self.process.copy()
        
        df = df[['Process','Stage','Output','Input','Display']]
        df = df[df['Stage']!= 'Sink']
        
        for idx in df.index:
            p = self.process_dict[idx] 
            
            ### Energy profile
            df.loc[p.idx,'energy_fuel'] = (p.NG_ratio+p.H2_ratio)*p.energy_intensity
            df.loc[p.idx,'energy_ele_clean'] = p.ele_ratio*p.clean_grid_ratio*p.energy_intensity
            df.loc[p.idx,'energy_ele_fossil'] = p.ele_ratio*(1-p.clean_grid_ratio)*p.energy_intensity
            
            
            ### Emission profile
            df.loc[p.idx,'CF_process'] = p.CF_process
            df.loc[p.idx,'CF_fuel'] = p.CF_heat
            df.loc[p.idx,'CF_ele'] = p.CF_ele
            df.loc[p.idx,'CF'] = p.CF_process+p.CF_heat+p.CF_ele
            df.loc[p.idx,'CF_ver'] = p.CF
            df.loc[p.idx,'yield'] = p.efficiency
      
        return df
    
    @property
    def CO2_table(self):
        df = self.process.copy()
        
        df = df[(df['CO2_oprt'] != 0) | (df.index == 'Use')]
        df = df[['Process','Stage','Output','Input','CO2_oprt','CO2_impt','CO2']]
        pos = df.columns.get_loc('Output') + 1
        df.insert(pos, 'CO2_heat', 0)
        df.insert(pos + 1, 'CO2_ele', 0)
        df.insert(pos + 2, 'CO2_proc', 0)
        df[['CO2_heat', 'CO2_ele', 'CO2_proc']] = df[['CO2_heat', 'CO2_ele', 'CO2_proc']].astype(float)
        
        df.loc['EoL','Stage'] = 'Secondary'
        df.loc['M','Stage'] = 'Primary'
        df.loc['E','Stage'] = 'Primary'
        df.loc['R1','Stage'] = 'Secondary'
        df.loc['R2','Stage'] = 'Secondary'
        df.loc['C1','Stage'] = 'Casting'
        df.loc['C2','Stage'] = 'Casting'
        df.loc[df['Stage'] == 'Fabrication','Stage'] = 'Fabricating'
        df.loc[df['Stage'] == 'Sink','Stage'] = 'End-Use'
        
        #for idx in ['F1','F2','F3','F4']:
            #df.loc[idx,'Stage'] = 'Deformation'
        #for idx in ['SC1','SC2','SC3']:
            #df.loc[idx,'Stage'] = 'Shape Casting'
            
        for idx in df.index:
            p = self.process_dict[idx]
            
            df.loc[idx,'CO2_heat'] = p.CO2_heat
            df.loc[idx,'CO2_ele'] = p.CO2_ele
            df.loc[idx,'CO2_proc'] = p.CO2_proc
            #df.loc[idx,'verify'] = abs(p.CO2_heat+p.CO2_ele+p.CO2_proc-p.CO2_oprt)>0.0001
            
            if idx == 'Use':
                df.loc['Use','CO2_impt'] = self.CO2_impt_use
                
        return df
    
    @property
    def total_CO2_table(self):
        df = self.CO2_table
        df.loc['Import','Stage'] = 'Import'
        df.loc['Import','CO2_oprt'] = df['CO2_impt'].sum()
        
        df = df.groupby('Stage')['CO2_oprt'].sum()
        df = df
        df = df.reset_index()
        
        df.rename(columns = {'CO2_oprt':'CO2'},inplace = True)
        df.set_index('Stage', inplace = True)
        df.sort_values(by = 'CO2', inplace = True, ascending = False)
        
        return df
    
    @property
    def dom_CO2_table(self):
        df = self.CO2_table
        df.drop(columns = ['Process','Output','Input'],inplace = True)
        df = df.groupby('Stage').sum()
        
        order = ['Primary','Secondary','Casting','Forming','Fabricating','End-Use']
        df = df.reindex(order)
        
        return df
    
    
    
    @property
    def import_CO2_table(self):
        df = self.process.copy()
        df = df[['Process','Stage','Input','Import','CO2_impt']]
        df = df[df['CO2_impt']!=0]
        df = df.groupby('Input').sum()
        df.rename(columns = {'CO2_impt':'CO2'}, inplace = True)
        df.sort_values(by = 'CO2', inplace = True, ascending = False)
                
        return df
    
    @property
    def trade_table(self):
        M = self.process_dict['M']
        C1 = self.process_dict['C1']
        C2 = self.process_dict['C2']
        
        process = self.process
        df = pd.DataFrame(columns = ['Produced','Import','Export'])
        df.loc['Alumina'] = [M.outflow - M.outflow_table.loc['W','amount']+M.export_amt,
                             self.E.import_amt, M.export_amt]
        df.loc['Primary\nMetal'] = [self.E.outflow - self.E.outflow_table.loc['W','amount']+self.E.export_amt,
                           self.R1.import_metal+self.R2.import_metal,
                           self.E.export_amt]
        df.loc['Wrought\nAlloys'] = [C1.outflow+C1.export_amt, 
                                    process[process['Input'] == 'Wrought Alloys']['Import'].sum(),
                                    C1.export_amt]
        df.loc['Foundry\nAlloys'] = [C2.outflow+C2.export_amt, 
                                    process[process['Input'] == 'Foundry Alloys']['Import'].sum(),
                                    C2.export_amt]
        df.loc['Semi'] = [self.semi_product_produced, self.semi_product_import, self.semi_product_export]
        df.loc['Products'] = [self.product_fabricated, self.product_import, self.product_export]
        df['dom_to_dom'] = df['Produced'] - df['Export']
        
        return df 

    @property
    def SC_profile(self):
        df = self.trade_table.copy()
        
        # Compute the new row first, but don't assign it yet
        new_row = df.loc["Wrought\nAlloys"] + df.loc["Foundry\nAlloys"]
        
        # Get the position of "Foundry Alloys"
        pos = df.index.get_loc("Wrought\nAlloys")
        
        # Drop the original two rows
        df = df.drop(["Wrought\nAlloys", "Foundry\nAlloys"])
        
        # Insert "Alloys" at the correct position
        df_upper = df.iloc[:pos]
        df_lower = df.iloc[pos:]
        df = pd.concat([df_upper, pd.DataFrame([new_row], index=["Alloys"]), df_lower])
        
        # scrap
        available_post = self.EoL.inflow
        available_pre = self.R1.internal_last + self.R2.internal_last
        df.loc['EoL_Scrap','Produced'] = available_post
        df.loc['EoL_Scrap','Import'] = 0.0 
        df.loc['EoL_Scrap','Export'] = self.EoL.export_amt
        df.loc['EoL_Scrap','dom_to_dom'] = available_post - self.EoL.export_amt
        
        df.loc['Pre_Scrap'] = available_pre
        df.loc['Pre_Scrap','Import'] = self.EoL.import_amt
        df.loc['Pre_Scrap','Export'] = 0.0 
        df.loc['Pre_Scrap','dom_to_dom'] = available_pre
        

        df['Consumed'] = df['Import'] + df['dom_to_dom']
        

        df['Import%'] = (df['Import']/df['Consumed']).apply(lambda x: f"{x * 100:.1f}%")
        df['Export%'] = (df['Export']/df['Produced']).apply(lambda x: f"{x * 100:.1f}%")
        
        
        return df
    
    
    
    @property
    def ctg_table(self):
        #Cradle-to-Gate: Embodied kgCO2e/kg output for each process output
        #this segment takes some time to execute, so by default muted by a "break"
        ctg_table = self.process.copy()
        ctg_table = ctg_table[['Process','Display']]
        ctg_table.rename(columns = {'Display':'Output'},inplace = True)
        ctg_table.drop(['Use', 'W'], inplace = True) 
        self.ctg_memo = {} # for dynamic programming ensuring each sector is only executed once
        for idx in ctg_table.index:
            CO2_ctg = self.process_dict[idx].CO2_ctg
            ctg_table.loc[idx,'kgCO2e/kg-Al'] = CO2_ctg
            self.ctg_memo[idx] = CO2_ctg
            
            
        ctg_table.loc['M','Type'] = 'Crude'
        ctg_table.loc['EoL','Type'] = 'Crude'
        ctg_table.loc['EoL','Output'] = 'Sorted EoL Scrap'
        ctg_table.loc['E','Type'] = 'Primary Metal'
        for idx in ['R1','R2']:
            ctg_table.loc[idx,'Type'] = 'Secondary Metal'
        for idx in ['C1','C2']:
            ctg_table.loc[idx,'Type'] = 'Alloy Ingots'
        for idx in ['F1','F2','F3','F4']:
            ctg_table.loc[idx,'Type'] = 'Deformations'
        for idx in ['SC1','SC2','SC3']:
            ctg_table.loc[idx,'Type'] = 'Shape Castings'
        for idx in ['P'+str(i) for i in range(1,9)]:
            ctg_table.loc[idx,'Type'] = 'Final Products'
            
        return ctg_table
    
    
    def update_A(self): 
        A = self.A.copy()
        
        R1_inflows = self.R1.inflow_table['amount']
        R2_inflows = self.R2.inflow_table['amount']
        R1_target = self.R1.target
        R2_target = self.R2.target
        
        for idx in R1_inflows.index:
            #A.loc[idx, 'R1'] = R1_inflows[idx]/R1_target
            A.loc[idx, 'R1'] = R1_inflows[idx]/self.R1.target
        for idx in R2_inflows.index:
            #A.loc[idx, 'R2'] = R2_inflows[idx]/R2_target
            A.loc[idx, 'R2'] = R2_inflows[idx]/self.R2.target
        
        if self.push_scrap == False:
            A.loc['E','R1'] = (1 - self.R1.scrap_ratio)*self.R1.dom_sweet_ratio
            A.loc['E','R2'] = (1 - self.R2.scrap_ratio)*self.R2.dom_sweet_ratio
        
        self.A = A
                                
        '''      
        outflow_matrix = self.outflow_matrix.copy()        
        outflow_matrix.drop(['Use', 'W'], inplace = True)
        A['R1'] = outflow_matrix['R1']
        A['R1'] = A['R1']/self.R1.target
        A['R2'] = outflow_matrix['R2']
        A['R2'] = A['R2']/self.R2.target
        A = A.astype(float) 
        self.A = A
        '''
        

    @property
    def flow(self):
        flow = self.flow_0.copy()
        
        for row in flow.index:
            source_idx = flow.loc[row, "source"]
            source = self.process_dict[source_idx]
            target_idx = flow.loc[row, "target"]
            
            flow.loc[row, "amount"] = source.outflow_table.loc[target_idx, "amount"]
            
        return flow
            
    @property
    def flow_full(self):
        df = self.flow.copy()
                
        # Part I: Non-products        
        for idx in self.process.index:
            ### Export flows of non-products 
            #export_amt = self.process.loc[idx, 'Export']
            stage = self.process.loc[idx, 'Stage']
            export_amt = self.process_dict[idx].export_amt
            
                        
            
            if export_amt > 0:
                row_idx = len(df.index)
                df.loc[row_idx,'source'] = idx
                
                df.loc[row_idx,'amount'] = export_amt
                df.loc[row_idx,'type'] = self.process.loc[idx, 'Output']

                if stage!= 'Source':
                    df.loc[row_idx,'target'] = 'Export_'+stage[:4]                    
                elif idx == 'M':
                    df.loc[row_idx,'target'] = 'Export_Oxide'
                elif idx == 'EoL':
                    df.loc[row_idx,'target'] = 'Export_Scrap'
       
            ### Import flows of non-products 
            #import_amt = self.process.loc[idx, 'Import']
            import_amt = self.process_dict[idx].import_amt
            stage = self.process.loc[idx, 'Stage']
            if import_amt > 0:
                row_idx = len(df.index)
                
                if idx == 'M':
                    df.loc[row_idx,'target'] = idx
                    df.loc[row_idx,'source'] = 'Import_Ore'
                    df.loc[row_idx, 'type'] = 'Bauxite'
                    df.loc[row_idx,'amount'] = import_amt

                elif idx == 'EoL':                    
                    df.loc[row_idx,"source"] = 'Import_EoL'
                    df.loc[row_idx,"target"] = 'EoL'
                    df.loc[row_idx, 'type'] = 'Scrap'
                    df.loc[row_idx,'amount'] = self.process_dict[idx].import_amt
                
                elif idx == 'E':
                    df.loc[row_idx,'source'] = 'Import_Oxide'
                    df.loc[row_idx,'target'] = idx
                    df.loc[row_idx, 'type'] = 'Alumina'
                    df.loc[row_idx,'amount'] = import_amt
                
                elif idx in ['R1','R2']:
                    df.loc[row_idx,'source'] = 'Import_Metal'
                    df.loc[row_idx,'target'] = idx
                    df.loc[row_idx, 'type'] = 'Aluminum_primary'                    
                    df.loc[row_idx,'amount'] = self.process_dict[idx].import_metal
                
                elif idx in ['C1','C2']:
                    df.loc[row_idx,'source'] = 'Import_Metal_C'
                    df.loc[row_idx,'target'] = idx                    
                    df.loc[row_idx,'type'] = self.process.loc['E','Output']
                    df.loc[row_idx,'amount'] = import_amt
                    
                elif stage == 'Forming':
                    df.loc[row_idx,'source'] = 'Import_'+stage[:4]
                    df.loc[row_idx,'target'] = idx
                                        
                    if idx in ['SC1','SC2','SC3']:
                        df.loc[row_idx,'type'] = 'Cast_ingot'
                        df.loc[row_idx,'amount'] = import_amt
                    else:
                        df.loc[row_idx,'type'] = 'Wrought_ingot'
                        df.loc[row_idx,'amount'] = import_amt
                
                elif stage == 'Fabrication':
                    inflow_table = self.process_dict[idx].inflow_table
                    forming_import = inflow_table[inflow_table.index.isin(['F1','F2','F3','F4'])]['import_amt'].sum()                            
                    casting_import = inflow_table[inflow_table.index.isin(['SC1','SC2','SC3'])]['import_amt'].sum()                            
                    
                    df.loc[row_idx,'source'] = 'Import_'+stage[:4]
                    df.loc[row_idx,'target'] = idx
                    df.loc[row_idx,'type'] = 'Forming'
                    df.loc[row_idx,'amount'] = forming_import
                    
                    row_idx = row_idx +1
                    df.loc[row_idx,'source'] = 'Import_'+stage[:4]
                    df.loc[row_idx,'target'] = idx
                    df.loc[row_idx,'type'] = 'Castings'
                    df.loc[row_idx,'amount'] = casting_import
                            
        # Part II: Products
        for idx in self.product_table.index:
            ### Export flows of products
            export_amt = self.product_table.loc[idx,'export']
            if export_amt > 0:
                row_idx = len(df.index)
                df.loc[row_idx,'source'] = idx
                df.loc[row_idx,'target'] = 'Export_Prod'
                df.loc[row_idx,'amount'] = export_amt
                df.loc[row_idx,'type'] = self.product_table.loc[idx, 'type']
                
                # if product exports are shown, need to adjust the outflow from P_idx to Use
                # after the updates, we still have:
                # self.process_dict['Use'].inflow = self.product_fabricated
                
                row_idx =  df[(df['source'] == idx) & (df['target'] == 'Use')].index                
                df.loc[row_idx,'amount'] = self.product_table.loc[idx, 'received'] - self.product_table.loc[idx, 'export']
            
            ### Import flows of products
            import_amt = self.product_table.loc[idx,'import']
            if import_amt > 0:
                row_idx = len(df.index)
                df.loc[row_idx,'source'] = 'Import_Prod'
                df.loc[row_idx,'target'] = 'Use'
                df.loc[row_idx,'amount'] = import_amt
                df.loc[row_idx,'type'] = self.product_table.loc[idx, 'type']
                
        ### Merge EoL
        row_idx = len(df.index)
        df.loc[row_idx,"source"] = 'EoL_0'
        df.loc[row_idx,"target"] = 'EoL'
        df.loc[row_idx,"amount"] = self.process_dict['EoL'].inflow
        df.loc[row_idx,"type"] = 'Scrap'
        
        ### Internal loops     
        for p_idx in ['SC1','SC2','SC3']:            
            row_idx = len(df.index)
            df.loc[row_idx,"source"] = p_idx
            df.loc[row_idx,"target"] = p_idx
            df.loc[row_idx,"amount"] = self.process_dict[p_idx].internal
            df.loc[row_idx,"type"] = 'Recycle'
        
        ### Hide Small Flows
        threshold = 0.01 if self.alternative_trade == False else 20
        
        for idx in df.index:
            if df.loc[idx,'amount'] == 0:
                #df.loc[idx,'amount'] = 0
                df.drop(idx, inplace = True)
                
            elif 'Import' in df.loc[idx,'source'] and df.loc[idx,'amount'] < threshold:
                df.drop(idx, inplace = True)
            elif 'Export' in df.loc[idx,'target'] and df.loc[idx,'amount'] < threshold:
                df.drop(idx, inplace = True)
                
                
        ### Waste flows
        for row_idx in df[df['target'] == 'W'].index:
        
            df.loc[row_idx, 'target'] = 'W_'+df.loc[row_idx, 'source']
                                               
        # On Sankey, the inflow of Use now becomes: self.domestic_consumed_product               
        return df