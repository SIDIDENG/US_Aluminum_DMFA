'''Metalscape'''
### Version 07/22/2025
### 07/22/2025: Breaking down CO2 into dom_to_dom, export, and import
### 06/20/2025: Added SC_profile dataframe
### 05/09/2025: Add alloy_demand_dict
### 03/11/2025: Adjusted time-series legend label related to AA DMFA
### 03/06/2025: Small adjustments to figures
### 03/01/2025: define properties primary_production and secondary production
### 12/13/2024: added SA argument
### 09/10/2024: added %change column to df_eva
### 09/09/2024: Adjusted Evaluation format (thousand separators for the specified rows)

from model import *
import matplotlib.pyplot as plt
import seaborn as sns
import matplotlib.patches as patches
from matplotlib.ticker import FuncFormatter

plt.style.use('seaborn-v0_8-ticks')
plt.rcParams["font.family"] = "Arial"
plt.rcParams['mathtext.default'] = 'regular'



#%%
warmup_period = 0 #4
sim_period = 0 #26 for 2050, 11 for 2035
   
sim = Simulation(warmup_period = warmup_period, sim_period = sim_period, 
                 yield_update = False, EoU_update = False, grid_update = False,
                 hydrogen = False, blue = False, PTC = False,
                 electrification = False,
                 technology_update = False,
                 alternative_trade = False, trade_mode = 'Island', #['Island','Surplus','Deficit']
                 CV = 0, lightweight = False,
                 SA = False, dom_alloy_ratio = 0.9, dom_semi_ratio = 0.9,
                 global_decarb_mode = 'Constant',
                 push_scrap = False,
                 record_process = True, display = True, stepwise = True, 
                 breakdown_emissions = False,
                 optimization = True, optimization_all = False)


#%%
### Overwrite system attributes
def reset_system():
    sim.yield_update = True
    sim.EoU_update = True
    sim.grid_update = True
    sim.hydrogen = True
    sim.electrification = True
    sim.technology_update = True   
    
#reset_system()


#%%
#reset_system()
sim.initialize()
sim.run()

#%%
process = sim.process
flow_0 = sim.flow_0
flow = sim.flow
flow_full = sim.flow_full
flow_full.to_csv("test_flow.csv", index = False)

Gamma = sim.Gamma
Waste = generate_consumption_target(sim.A, sim.Gamma, sim.Sink)[1]['W']


Sink = sim.Sink
parameter = g.parameter
EoL_table = g.EoL_table
CF_ele_table = g.CF_ele_table
CF_NREL_table = g.CF_NREL_table
#consumption_target_0 = g.consumption_target.
#consumption_target = sim.consumption_target
product_trade_table = g.product_trade_table
outflow_matrix = sim.outflow_matrix
#net_flow_matrix = sim.net_flow_matrix
#IO_matrix_0 = g.IO_matrix
IO_matrix= sim.IO_matrix
semi_CF = g.semi_CF
A = sim.A
A0 = g.A
if sim.optimization == True:
    scrap_supply_old = g.scrap_supply_old
    alloy_demand_dict = sim.alloy_demand_dict
    alloy_supply_dict = sim.alloy_supply_dict
    scrap_supply = sim.scrap_supply


fraction_matrix = sim.fraction_matrix
fraction_matrix_0 = sim.fraction_matrix_0
fraction_matrix_1 = sim.generate_fraction_matrix()
total_CO2 = sim.total_CO2

    
process_dict = sim.process_dict
system_track = sim.system_track
system_deets = sim.system_deets
sim_track = sim.sim_track
# In[]

print("  inflow outflow")
for p in sim.process_list:
    print("{} {:.2f} {:.2f} {}".format(p.idx,
                 p.inflow_0+p.import_amt - p.export_amt,
                 p.outflow,
                 abs(p.inflow_0+p.import_amt - p.export_amt-p.outflow)<0.000001))
print("\n exp_dmd exp")    
for p in sim.process_list:
    if p.idx not in ['Use','W']:
        print("{} {:.2f} {:.2f}".format(p.idx, 
                                           Sink.loc[p.idx, 'Offset']+Sink.loc[p.idx, 'Semi'],
                                           p.export_amt))
    
# real and Leontief targets
targets = sim.targets

end_use = sim.end_use
product_table = sim.product_table
scrap_allocation = sim.scrap_allocation


M = process_dict['M']
EoL = process_dict['EoL']
E = process_dict['E']
R1 = process_dict['R1']
R2 = process_dict['R2']
F1 = process_dict['F1']
F2 = process_dict['F2']
F3 = process_dict['F3']
F4 = process_dict['F4']
SC1  = process_dict['SC1']
SC2  = process_dict['SC2']
SC3  = process_dict['SC3']
Use = process_dict['Use']
W = process_dict['W']
C1 = process_dict['C1']
C2 = process_dict['C2']
for i in range(1,9):
    idx = 'P'+str(i)
    globals()[idx] = process_dict[idx]
    



print()
print("Exogenous input: ", sim.total_input)
print('total import: ', sim.total_import)
print('Internal scrap intake: ', sim.internal_scrap_intake)
'''
if sim.env.now>2:
    print("Last year recycled scrap: ", sim.system_track.loc[sim.env.now - 2, 'recycled'] , sim.internal_scrap_intake)
if sim.env.now == 2:
    print("Last year recycled scrap: ", sim.total_recycled_scrap_0, sim.internal_scrap_intake)
if sim.env.now == 1:
    print("Recycled scrap: ", sim.system_track.loc[0, 'recycled'], sim.internal_scrap_intake)
'''
print("------------------------------------------")


print("System inflow: ", sim.system_inflow)
print("System inflow: ", sim.total_input+sim.total_import+sim.internal_scrap_intake)

print()
print("Domestic product: ", sim.domestic_consumed_product)
print("Internal scrap generated: ", sim.internal_scrap_generated)
print("total waste: ", sim.total_waste)
print('total export: ', sim.total_export)
print("------------------------------------------")
print("system outflow: ", sim.system_outflow)
print("system outflow: ", sim.domestic_consumed_product+sim.internal_scrap_generated+sim.total_waste+sim.total_export)


print()
print()


print("material efficiency: ", sim.material_efficiency)
print()
print("CO2: ", sim.total_CO2, "MT CO2e")
print("Carbon Intensity: ", sim.carbon_intensity)

print()
if sim.record_process == True:
    print(process.loc[["M","EoL"],["Target","Inflow_0"]])
print("Source demand: ", sim.source_demand)
print("Source input: ", sim.total_input)

# In[Examine process]
p = process_dict['P1']
print()
print()
print(p.idx, p.name)

outflow_table = p.outflow_table
inflow_table = p.inflow_table
process_track = p.process_track
#ctg = p.ctg



print("inflow: ",p.inflow_0)
print("import: ", p.import_amt)
print("input: ", p.inflow_0+p.import_amt)
print("----------")
print("outflow: ", p.outflow)
print("export: ", p.export_amt)
print("output: ", p.outflow+p.export_amt)
print()

# In[Overall Summary]    

# fraction matrix
summation = 0
for row in fraction_matrix.index:
    summation = summation + fraction_matrix.loc[row].sum()
print('fraction matrix row sums = 1: ', abs(summation/fraction_matrix.shape[0] - 1)<0.00001)    

# quick verficiation
print("total flow target verification: ", round(sim.consumption_target['Target'].sum()))
print("Inflow = Outflow? ",abs(sim.system_inflow - sim.system_outflow)<0.000001)


# In[Evaluating Performance Metrics]

def Evaluate():
    '''
    Compare benchmarking metrics for 2020, 2035, and 2050
    
        Only execute if sim.t>30
    
    Returns
    -------
    df_eva : Dataframe
        This dataframe will be saved as a workbook

    '''
    df_eva = system_track[["year","consumed","primary","secondary",
                           "EoL","EoL_recycled",
                           "CO2"]]
    
    if len(df_eva) > 16:
        year_list = [0, 15, warmup_period + sim_period]
    elif warmup_period + sim_period> 0:
        year_list = [0, warmup_period + sim_period]
    else:
        year_list = [0]
    
    df_eva = df_eva[df_eva.index.isin(year_list)]
    df_eva = df_eva.transpose()
    
    df_eva['change'] = (df_eva[warmup_period + sim_period] - df_eva[0]) / df_eva[0]
    
    for row in df_eva.index:
        if row in ["year", "consumed", "primary", "secondary", "EoL", "EoL_recycled", "pre_recycled"]:
            df_eva.loc[row, year_list] = df_eva.loc[row, year_list].apply(round)
            df_eva.loc[row, year_list] = df_eva.loc[row, year_list].apply(lambda x: f"{int(x):,}")
    
        elif row in ["material_efficiency", "alloy_efficiency"]:
            df_eva.loc[row, year_list] = df_eva.loc[row, year_list].apply(lambda x: f"{x * 100:.1f}%")
        else:
            df_eva.loc[row, year_list] = df_eva.loc[row, year_list].apply(lambda x: f"{x:.1f}")
    
    # Apply the percentage format to the 'change' column
    
    df_eva['change'] = df_eva['change'].apply(lambda x: "0.0%" if abs(x) < 0.0001 else f"{x * 100:.1f}%")
    
    
    # Save to Excel
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    excel_path = "Evaluation.xlsx"
    df_eva.to_excel(excel_path, index = True)
    
    
    # Apply formatting
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Set font to Times New Roman and align text to center
    font = Font(name="Times New Roman", size = 14)
    alignment = Alignment(horizontal="right", vertical="center")
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = font
            cell.alignment = alignment

    wb.save(excel_path)
    return df_eva


df_eva = Evaluate()
SC_profile = sim.SC_profile
SC_profile.to_excel('SC_profile.xlsx', index= True)

# In[Stats]

def show_SC_profile():
    print('Idx, Import%','Export')
    for col in A.columns:
        
        print(f'{col} {A[col].sum():.2f} {Sink.loc[col,'Semi'] +Sink.loc[col,'Offset']:.2f}kt')

show_SC_profile()

# In[Test bed]
# Time-series Plots
import matplotlib.ticker as ticker

df_yx = g.df_yx

def plot_scrap_series(system_track, t,
                      top_scale = 1.2):
    
    x = system_track['year']
    
    plt.figure(figsize = (15,10), dpi = 300)
    
    ### AA DMFA data
    plt.plot(x,system_track['consumed'], color = 'darkred', label = 'Domestic consumption of finished products (AA-informed DMFA, as simulation input)')
    
    plt.plot(x,system_track['old_post_0'], color = 'gray', label = r'$H(t)$: Collected scrap from products consumed before 2020 (AA-informed DMFA, as simulation input)',
             linestyle = '--', linewidth = 2)
    if t == 31:
        plt.plot(x,df_yx["Collected"], color = 'green', label = 'Collected post-consumer scrap (AA-informed DMFA, for reference)',
                 linewidth = 10, alpha = 0.5)
    
    ### Simulation output
    plt.plot(x,system_track['new_post'], color = 'teal', label = 'Collected scrap from products consumed after 2020 (simulation output)',
             linewidth = 2)
    plt.plot(x,system_track["EoL_collected"], color = 'black', label = r'$X_S(t)$: Collected post-consumer scrap (simulation output)',
             linewidth = 2) #
    
        
    if sim.t>2:
        plt.xlim(left = 2021, right = max(x))
    plt.ylim(bottom = 0, top = top_scale*max(system_track['consumed']))
    
    if t > 30:
        plt.xticks([2021, 2025, 2030, 2035, 2040,2045, 2050], fontsize=18)
    plt.gca().yaxis.set_major_formatter(ticker.FuncFormatter(lambda x, pos: f'{int(x):,}'))
    plt.yticks(fontsize = 18)
    
    
    plt.xlabel("Year (n)", fontsize = 18)
    plt.ylabel("Mass (kt)", fontsize = 18)
    
    plt.legend(loc = 2, fontsize = 18, framealpha = 0)
    
    plt.tight_layout()    
    plt.show()
    
    
plot_scrap_series(system_track, sim.t, top_scale = 1.25)

# In['CO2 breakdown]
if sim.breakdown_emissions == True:
    df_CO2 = sim.system_track[['year','import_to_dom_CO2','dom_to_dom_CO2','export_CO2']]
    df_CO2.set_index('year',inplace = True)

def CO2_plot(df):
    # Define data structure
    x = np.arange(2020, 2051)
    bottom = np.zeros(len(df))
    f_size = 16
    label_map = {'import_to_dom_CO2':'Import emissions attributable to domestic consumption', 
                  'dom_to_dom_CO2': 'Domestic emissions attributable to domestic consumption',
                  'export_CO2': 'Emissions (import + domestic) attributable to exports'}
    color_map = {'import_to_dom_CO2':'red', 
                  'dom_to_dom_CO2': 'blue',
                  'export_CO2': 'lightskyblue'}
    
    plt.figure(figsize = (12, 8), dpi = 300)
    
    # plot and fill the area
    for col in df.columns:
        y = df[col].to_numpy()
        
        
        plt.plot(x, bottom+y, color = color_map[col], alpha = 0.5)
        plt.fill_between(x, bottom, bottom+y, color = color_map[col],
                         alpha=0.2, label= label_map[col])
        
        # Annotate the region with the index label        
        #mid_y = bottom[len(x)-1] + y[len(x)-1]/2  # Find the middle of the filled area
        #plt.text(2050, mid_y, col, fontsize = f_size, color='black', ha='center', va='center')
        
        
        # Update bottom
        bottom = bottom + y
        
    # set axis limits
    plt.xlim(left = min(x), right = max(x))
    ax1 = plt.gca()
    ax1.set_ylim(bottom = 0, top = 80)
    
    # set axis labels
    plt.xlabel("Year", fontsize = f_size)
    plt.ylabel("Annual GHG Emissions (Mt $CO_2e$/year)", fontsize = f_size)
    
    # set axis ticks
    
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{x:,.0f}'))
    ax1.tick_params(axis='both', labelsize = f_size)
    
    
    # Add mirrored y-axis on the right
    
    ax2 = ax1.twinx()
    
    # Match y-limits and ticks
    ax2.set_ylim(ax1.get_ylim())
    ax2.set_yticks(ax1.get_yticks())
    ax2.set_yticklabels([f'{tick:,.0f}' for tick in ax1.get_yticks()])
    ax2.tick_params(axis='y', labelsize=f_size)

    
    ax1.legend(fontsize = f_size, facecolor='white', frameon = True)

    
        
    
    plt.show()
    
    
 

def test_breakdown():
    total_1 = 0
    total_2 = 0
    for p in sim.process_list:    
        if p.idx not in ['Use','W']:
            total_1 += p.CO2
            total_2 += p.CO2_export + p.CO2_import_to_dom + p.CO2_dom_to_dom
            print(f'{p.idx} {p.CO2:.2f} {p.CO2_export + p.CO2_import_to_dom + p.CO2_dom_to_dom:.2f}') 
        elif p.idx == 'Use':
            total_1 += p.CO2
            total_2 += p.CO2_import

    print(f'{total_1:.2f} {total_2:.2f}')
    
if sim.t == 31 and sim.breakdown_emissions == True:
    CO2_plot(df_CO2) 
    test_breakdown()

# In[]

def monitor_deets(df = system_deets, metrics = None, mode = 'single',
                  y_label = None, global_num_format = None, 
                  legend_loc = 'lower right', y_max_scale = 1.1):    
    
    num_metric = len(metrics)        
    x = system_deets['year']
    y_max = df[metrics].max().max()
    y_min = df[metrics].min().min()
    
    f_size = 16    
    
    if mode == 'single':
        plt.figure(figsize = (14,8), dpi = 300)
        ax = plt.gca()
        for metric in metrics:
            y = system_deets[metric]
            plt.plot(x, y,
                     label = df_config.loc[metric,'label'], color = df_config.loc[metric,'color'])
        
        #Axis labels
        ax.set_xlabel("Year", fontsize = f_size)
        
        # Primary Y-axis (Left)
        ax.set_xlim(left = min(x), right = max(x))
        ax.set_ylim(bottom = 0, top = 1.1*y_max)
        ax.set_xticks(np.arange(min(x), max(x)+1,5))
        ax.tick_params(axis = 'both', labelsize = f_size)
        set_ytick_format(ax, global_num_format)
        
        # Make twin Y-axis identical        
        ax_twin = ax.twinx()
        ax_twin.set_ylim(ax.get_ylim())
        ax_twin.tick_params(axis = 'both', labelsize = f_size)
        set_ytick_format(ax_twin, global_num_format)
        
        ### add legend            
        ax.legend(fontsize = f_size, facecolor='none', edgecolor='none') 
        
    if mode == 'subplots':
        fig, axs = plt.subplots(num_metric,1,figsize=(14, 8), dpi=300)
                
        for metric, ax in zip(metrics,axs):
            
            y = system_deets[metric]
            num_format = df_config.loc[metric,'format']
            
            ### Plot            
            ax.plot(x, y,
                    label = df_config.loc[metric,'label'], color = df_config.loc[metric,'color'])
            
            # Axis labels
            axs[-1].set_xlabel("Year", fontsize = f_size)
            
            ### lims and ticks
            ax.set_xlim(left = min(x), right = max(x)) 
            ax.set_ylim(bottom = 0, top = y_max_scale*max(y))
            ax.set_xticks(np.arange(min(x), max(x)+1, 5))            
            ax.tick_params(axis = 'both', labelsize = f_size)
            
            set_ytick_format(ax, num_format)
            
            ### add legend            
            ax.legend(fontsize = f_size,
                      loc = legend_loc, facecolor='none', edgecolor='none')  #bbox_to_anchor=(1.0, 0)
              
    
    plt.show()


df_config = pd.DataFrame(columns = ['metric', 'label','color', 'format'])
df_config.set_index('metric', inplace = True)
df_config.loc['gCO2_per_MJ_ele'] = ['g $CO_2e$ per MJ electricity delivered', 'steelblue', 'decimal']
df_config.loc['clean_grid_ratio'] = ['Clean and renewable electricity%', 'black', 'percentage']
df_config.loc['inert_anode_ratio'] = ['Inert anode ratio (%)','cyan','percentage']
df_config.loc['E_epsilon'] = ['U.S. Electrolysis Energy intensity ($\epsilon_E$, MJ-del/kg-in)','palevioletred', 'decimal']
df_config.loc['E_CF'] = ['U.S. Electrolysis Carbon intensity ($C_E$, kg $CO_2e$/MJ-del)', 'darkgoldenrod','decimal']


def set_ytick_format(ax, num_format):
### y-axis format
    if num_format == 'integer':                
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(x)}'))
    elif num_format == 'decimal':
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{x:.1f}"))
    elif num_format == 'percentage':
        ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f"{x * 100:.0f}%"))
    
if sim.t>1:
    
    monitor_deets(system_deets, ['gCO2_per_MJ_ele', 'gCO2_per_MJ_ele'], mode = 'single', 
                  y_label = ['g $CO_2e$/MJ-del'], global_num_format = 'decimal')
    monitor_deets(system_deets, ['gCO2_per_MJ_ele', 'clean_grid_ratio'], mode = 'subplots')
    
    monitor_deets(system_deets, ['inert_anode_ratio', 'E_epsilon', 'E_CF'], mode = 'subplots',
                  legend_loc = 'best', y_max_scale = 1.2)
    
    
    
    

